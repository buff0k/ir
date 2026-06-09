# Copyright (c) 2026, BuFf0k and contributors
# For license information, please see license.txt

from __future__ import annotations

import base64
import json
from io import BytesIO

import frappe
from frappe.utils import add_days, getdate, nowdate


INDUCTION_TYPE_FIELD_MAP = {
    "Training": "is_training",
    "Licence": "is_licence",
    "Qualification": "is_qualification",
    "Authorisation": "is_authorisation",
}


@frappe.whitelist()
def download_training_matrix_excel(filters=None):
    """Return a base64 XLSX export for the currently filtered Training Matrix.

    The report itself keeps the full JSON payload used by the card formatter.
    This export intentionally writes only the expiry date for each dynamic
    induction column.
    """
    filters = _coerce_filters(filters)
    columns, data = execute(filters)

    try:
        from openpyxl import Workbook
        from openpyxl.formatting.rule import FormulaRule
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception:
        frappe.throw(
            "The Python package openpyxl is required to export the Training Matrix to Excel."
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Training Matrix"

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    expired_fill = PatternFill("solid", fgColor="F4CCCC")
    expiring_fill = PatternFill("solid", fgColor="FFF2CC")

    export_columns = [c for c in columns if c.get("fieldname")]

    for col_idx, col in enumerate(export_columns, start=1):
        cell = ws.cell(
            row=1,
            column=col_idx,
            value=col.get("label") or col.get("fieldname"),
        )
        cell.font = Font(bold=True)
        cell.fill = header_fill

    induction_col_indexes = []

    for row_idx, row in enumerate(data, start=2):
        for col_idx, col in enumerate(export_columns, start=1):
            fieldname = col.get("fieldname")
            value = row.get(fieldname)

            if fieldname and fieldname.startswith("ind_"):
                if col_idx not in induction_col_indexes:
                    induction_col_indexes.append(col_idx)

                expiry = _extract_expiry(value)
                if expiry:
                    cell = ws.cell(row=row_idx, column=col_idx, value=getdate(expiry))
                    cell.number_format = "yyyy-mm-dd"
                else:
                    ws.cell(row=row_idx, column=col_idx, value=None)

                continue

            ws.cell(row=row_idx, column=col_idx, value=value)

    # Freeze header + Tracking, Employee, Employee Name, Employee ID Number
    ws.freeze_panes = "E2"
    ws.auto_filter.ref = ws.dimensions

    for col_idx, col in enumerate(export_columns, start=1):
        width = int((col.get("width") or 120) / 7)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(width, 38))

    last_row = max(ws.max_row, 2)

    for col_idx in induction_col_indexes:
        col_letter = get_column_letter(col_idx)
        cell_ref = f"{col_letter}2"
        range_ref = f"{col_letter}2:{col_letter}{last_row}"

        ws.conditional_formatting.add(
            range_ref,
            FormulaRule(
                formula=[f"AND(ISNUMBER({cell_ref}),{cell_ref}<TODAY())"],
                fill=expired_fill,
            ),
        )

        ws.conditional_formatting.add(
            range_ref,
            FormulaRule(
                formula=[
                    f"AND(ISNUMBER({cell_ref}),{cell_ref}>=TODAY(),{cell_ref}<=TODAY()+30)"
                ],
                fill=expiring_fill,
            ),
        )

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    filename = f"training_matrix_{nowdate()}.xlsx"

    return {
        "filename": filename,
        "content": base64.b64encode(out.read()).decode("utf-8"),
        "type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }


def execute(filters=None):
    filters = _coerce_filters(filters)

    today = getdate(nowdate())
    warn_date = getdate(add_days(today, 90))

    # Resolve branch scope ONCE (supports direct Branch filter and Area Setup -> branches)
    branches = _resolve_branch_set(filters)

    # If a branch scope is provided, we must scope BOTH:
    #  - Tracking docs
    #  - Employees list
    employees, tracking_docs_all = _get_employees_and_tracking_scoped(filters, branches)

    if not employees:
        return _base_columns(), []

    emp_ids = [e["name"] for e in employees]
    emp_map = {e["name"]: e for e in employees}

    # Ensure tracking docs are limited to these employees
    tracking_docs_all = [t for t in tracking_docs_all if t.get("employee") in emp_map]

    # Apply designation filter:
    # effective_designation = tracking.designation if set else employee.designation
    designation_filter = (filters.get("designation") or "").strip() or None
    tracking_docs = _filter_tracking_by_designation(
        tracking_docs_all,
        emp_map,
        designation_filter,
    )

    # Determine final included employees:
    # - If no designation filter: employees already scoped by branch
    # - If designation filter: include employee if employee.designation matches
    #   OR has a matching tracking doc
    included_employees = _filter_employees_by_designation_fallback(
        employees,
        tracking_docs,
        designation_filter,
    )

    if not included_employees:
        return _base_columns(), []

    included_emp_ids = [e["name"] for e in included_employees]
    included_emp_map = {e["name"]: e for e in included_employees}

    # Re-filter tracking docs to included employees only
    tracking_docs = [t for t in tracking_docs if t["employee"] in included_emp_map]

    # Build competency columns from union of required inductions across tracking docs in scope
    tracking_names = [t["name"] for t in tracking_docs]
    required_by_tracking, all_inductions = _get_required_inductions(tracking_names)

    induction_meta = _get_induction_meta(all_inductions) if all_inductions else {}

    all_inductions = _filter_inductions_by_type(
        all_inductions,
        induction_meta,
        filters.get("induction_type"),
    )

    required_by_tracking = _filter_required_by_allowed_inductions(
        required_by_tracking,
        all_inductions,
    )

    induction_name_map = {
        name: (meta.get("training_name") or name)
        for name, meta in induction_meta.items()
    }

    if all_inductions:
        comp_columns, comp_field_map, ordered_inductions = _build_competency_columns(
            all_inductions,
            induction_name_map,
        )
    else:
        comp_columns, comp_field_map, ordered_inductions = [], {}, []

    # Index records for included employees
    record_index = _index_records(included_emp_ids, today)

    # Output rows:
    # - one row per tracking doc for that employee
    # - else one blank employee row
    data = []
    tracking_by_employee = {}

    for t in tracking_docs:
        tracking_by_employee.setdefault(t["employee"], []).append(t)

    for emp in included_employees:
        emp_id = emp["name"]
        t_list = tracking_by_employee.get(emp_id, [])

        if t_list:
            for t in sorted(
                t_list,
                key=lambda x: ((x.get("branch") or ""), (x.get("name") or "")),
            ):
                row = _base_row_tracking(t, included_emp_map)
                required = set(required_by_tracking.get(t["name"], []))

                for induction_id in ordered_inductions:
                    fieldname = comp_field_map[induction_id]

                    if induction_id not in required:
                        row[fieldname] = ""
                        continue

                    payload = _build_cell_payload(
                        emp=emp_id,
                        induction=induction_id,
                        idx=record_index,
                        today=today,
                        warn_date=warn_date,
                    )

                    row[fieldname] = json.dumps(payload)

                data.append(row)
        else:
            row = _base_row_employee_only(emp)

            for induction_id in ordered_inductions:
                row[comp_field_map[induction_id]] = ""

            data.append(row)

    columns = _base_columns() + comp_columns
    return columns, data


# -----------------------
# General helpers
# -----------------------

def _coerce_filters(filters):
    if not filters:
        return {}

    if isinstance(filters, str):
        try:
            return json.loads(filters)
        except Exception:
            return {}

    return filters


def _extract_expiry(value):
    if not value:
        return None

    try:
        obj = json.loads(value) if isinstance(value, str) else value
    except Exception:
        return None

    if not isinstance(obj, dict):
        return None

    return obj.get("expiry")


# -----------------------
# Columns / Rows
# -----------------------

def _base_columns():
    return [
        {
            "fieldname": "tracking",
            "label": "Tracking",
            "fieldtype": "Link",
            "options": "Employee Induction Tracking",
            "width": 170,
        },
        {
            "fieldname": "employee",
            "label": "Employee",
            "fieldtype": "Link",
            "options": "Employee",
            "width": 120,
        },
        {
            "fieldname": "employee_name",
            "label": "Employee Name",
            "fieldtype": "Data",
            "width": 220,
        },
        {
            "fieldname": "employee_id_number",
            "label": "Employee ID Number",
            "fieldtype": "Data",
            "width": 160,
        },
        {
            "fieldname": "branch",
            "label": "Branch",
            "fieldtype": "Link",
            "options": "Branch",
            "width": 140,
        },
        {
            "fieldname": "designation",
            "label": "Designation",
            "fieldtype": "Link",
            "options": "Designation",
            "width": 160,
        },
    ]


def _base_row_tracking(tracking_row, emp_map):
    emp = emp_map.get(tracking_row["employee"]) or {}
    effective_designation = tracking_row.get("designation") or emp.get("designation")

    return {
        "tracking": tracking_row["name"],
        "employee": tracking_row["employee"],
        "employee_name": emp.get("employee_name"),
        "employee_id_number": emp.get("za_id_number"),
        "branch": tracking_row.get("branch"),
        "designation": effective_designation,
    }


def _base_row_employee_only(emp_row):
    return {
        "tracking": None,
        "employee": emp_row["name"],
        "employee_name": emp_row.get("employee_name"),
        "employee_id_number": emp_row.get("za_id_number"),
        "branch": emp_row.get("branch"),
        "designation": emp_row.get("designation"),
    }


# -----------------------
# Branch scope helpers
# -----------------------

def _resolve_branch_set(filters):
    branch = (filters.get("branch") or "").strip() or None
    area_setup = (filters.get("area_setup") or "").strip() or None

    branches = set()

    if area_setup:
        rows = frappe.get_all(
            "Branch Selector",
            filters={
                "parenttype": "Area Setup",
                "parent": area_setup,
            },
            fields=["branch"],
            limit_page_length=10000,
        )

        for r in rows:
            if r.get("branch"):
                branches.add(r["branch"])

    if branch:
        branches.add(branch)

    return sorted(branches)


def _get_employees_and_tracking_scoped(filters, branches):
    """
    If branches are provided, ensure BOTH employees and tracking docs are scoped
    to those branches.

    Employee scope when branches are provided:
      - employees whose Employee.branch is in branches
      - OR employees who have at least one tracking doc in branches

    Also applies employee_status and explicit employee filter.
    """
    employee = (filters.get("employee") or "").strip() or None
    employee_status = (filters.get("employee_status") or "Active").strip() or "Active"

    emp_filters_base = {}

    if employee_status != "All":
        emp_filters_base["status"] = employee_status

    # If no branch scope, behave like before:
    # employees first, then tracking docs scoped by those employees.
    if not branches:
        employees = _get_employees(filters)

        if not employees:
            return [], []

        emp_ids = [e["name"] for e in employees]
        tracking_docs_all = _get_tracking_docs(emp_ids, branches=None)

        return employees, tracking_docs_all

    # Branch scope exists:
    # 1) Pull tracking docs in those branches, and employee if explicitly selected.
    tracking_filters = {}

    if employee:
        tracking_filters["employee"] = employee

    tracking_docs_all = _get_tracking_docs(
        employee_ids=None,
        branches=branches,
        extra_filters=tracking_filters,
    )

    emp_ids_from_tracking = {
        t["employee"]
        for t in tracking_docs_all
        if t.get("employee")
    }

    # 2) Pull employees whose Employee.branch is in branches.
    emp_filters_branch = dict(emp_filters_base)
    emp_filters_branch["branch"] = ["in", branches]

    if employee:
        emp_filters_branch["name"] = employee

    employees_branch = frappe.get_all(
        "Employee",
        filters=emp_filters_branch,
        fields=[
            "name",
            "employee_name",
            "za_id_number",
            "designation",
            "branch",
            "status",
        ],
        order_by="employee_name asc",
        limit_page_length=5000,
    )

    # 3) Pull employees that appear in tracking docs,
    # even if their Employee.branch does not match.
    employees_tracking = []

    if emp_ids_from_tracking:
        emp_filters_tracking = dict(emp_filters_base)
        emp_filters_tracking["name"] = ["in", sorted(emp_ids_from_tracking)]

        if employee:
            emp_filters_tracking["name"] = employee

        employees_tracking = frappe.get_all(
            "Employee",
            filters=emp_filters_tracking,
            fields=[
                "name",
                "employee_name",
                "za_id_number",
                "designation",
                "branch",
                "status",
            ],
            order_by="employee_name asc",
            limit_page_length=5000,
        )

    # Merge unique by Employee.name
    merged = {}

    for e in employees_branch:
        merged[e["name"]] = e

    for e in employees_tracking:
        merged[e["name"]] = e

    employees = sorted(
        merged.values(),
        key=lambda x: (x.get("employee_name") or "").lower(),
    )

    return employees, tracking_docs_all


# -----------------------
# Employees - no branch scope
# -----------------------

def _get_employees(filters):
    employee = (filters.get("employee") or "").strip() or None
    employee_status = (filters.get("employee_status") or "Active").strip() or "Active"

    emp_filters = {}

    if employee:
        emp_filters["name"] = employee

    if employee_status != "All":
        emp_filters["status"] = employee_status

    return frappe.get_all(
        "Employee",
        filters=emp_filters,
        fields=[
            "name",
            "employee_name",
            "za_id_number",
            "designation",
            "branch",
            "status",
        ],
        order_by="employee_name asc",
        limit_page_length=2000,
    )


# -----------------------
# Tracking Docs
# -----------------------

def _get_tracking_docs(employee_ids=None, branches=None, extra_filters=None):
    """
    Flexible tracking getter:
      - If employee_ids is provided: restrict employee in employee_ids
      - If branches is provided: restrict branch in branches
      - extra_filters: dict merged in, e.g. {"employee": "EMP-0001"}
    """
    t_filters = {}

    if employee_ids:
        t_filters["employee"] = ["in", employee_ids]

    if branches:
        t_filters["branch"] = ["in", branches]

    if extra_filters:
        t_filters.update(extra_filters)

    return frappe.get_all(
        "Employee Induction Tracking",
        filters=t_filters,
        fields=[
            "name",
            "employee",
            "branch",
            "designation",
        ],
        order_by="employee asc, branch asc, modified desc",
        limit_page_length=20000,
    )


def _filter_tracking_by_designation(tracking_docs, emp_map, designation_filter):
    if not designation_filter:
        return tracking_docs

    out = []

    for t in tracking_docs:
        emp = emp_map.get(t["employee"]) or {}
        effective_designation = t.get("designation") or emp.get("designation")

        if effective_designation == designation_filter:
            out.append(t)

    return out


def _filter_employees_by_designation_fallback(
    employees,
    tracking_docs,
    designation_filter,
):
    if not designation_filter:
        return employees

    emp_with_matching_tracking = {
        t["employee"]
        for t in tracking_docs
    }

    out = []

    for e in employees:
        if e.get("designation") == designation_filter or e["name"] in emp_with_matching_tracking:
            out.append(e)

    return out


# -----------------------
# Required inductions per tracking doc
# -----------------------

def _get_required_inductions(tracking_names):
    tracking_names = [t for t in tracking_names if t]

    if not tracking_names:
        return {}, []

    children = frappe.get_all(
        "Employee Required Inductions",
        filters={
            "parent": ["in", tracking_names],
            "parenttype": "Employee Induction Tracking",
        },
        fields=[
            "parent",
            "induction",
        ],
        limit_page_length=200000,
    )

    required_by_tracking = {}
    all_inductions = set()

    for c in children:
        parent = c.get("parent")
        ind = c.get("induction")

        if not parent or not ind:
            continue

        required_by_tracking.setdefault(parent, []).append(ind)
        all_inductions.add(ind)

    return required_by_tracking, list(all_inductions)


def _get_induction_meta(induction_ids):
    if not induction_ids:
        return {}

    rows = frappe.get_all(
        "Employee Induction",
        filters={
            "name": ["in", induction_ids],
        },
        fields=[
            "name",
            "training_name",
            "is_training",
            "is_licence",
            "is_qualification",
            "is_authorisation",
        ],
        limit_page_length=10000,
    )

    return {
        r["name"]: r
        for r in rows
    }


def _filter_inductions_by_type(induction_ids, induction_meta, induction_type):
    induction_type = (induction_type or "").strip()

    if not induction_type:
        return list(induction_ids)

    type_field = INDUCTION_TYPE_FIELD_MAP.get(induction_type)

    if not type_field:
        return list(induction_ids)

    return [
        ind
        for ind in induction_ids
        if (induction_meta.get(ind) or {}).get(type_field)
    ]


def _filter_required_by_allowed_inductions(required_by_tracking, allowed_inductions):
    allowed = set(allowed_inductions or [])

    if not allowed:
        return {
            parent: []
            for parent in required_by_tracking
        }

    return {
        parent: [
            ind
            for ind in inductions
            if ind in allowed
        ]
        for parent, inductions in required_by_tracking.items()
    }


def _build_competency_columns(induction_ids, induction_name_map):
    if not induction_ids:
        return [], {}, []

    ordered = sorted(
        induction_ids,
        key=lambda x: (induction_name_map.get(x) or x).lower(),
    )

    columns = []
    field_map = {}

    card_col_width = 260

    for ind in ordered:
        fieldname = f"ind_{frappe.scrub(ind)}"
        field_map[ind] = fieldname

        columns.append(
            {
                "fieldname": fieldname,
                "label": induction_name_map.get(ind) or ind,
                "fieldtype": "Data",
                "width": card_col_width,
            }
        )

    return columns, field_map, ordered


# -----------------------
# Records index - employee + training
# -----------------------

def _index_records(employee_ids, today):
    rows = frappe.get_all(
        "Employee Induction Record",
        filters={
            "employee": ["in", employee_ids],
        },
        fields=[
            "name",
            "employee",
            "training",
            "training_date",
            "valid_to",
            "docstatus",
        ],
        order_by="training_date desc",
        limit_page_length=200000,
    )

    idx = {}

    for r in rows:
        emp = r.get("employee")
        trn = r.get("training")

        if not emp or not trn:
            continue

        key = (emp, trn)

        if key not in idx:
            idx[key] = {
                "submitted": None,
                "scheduled": None,
            }

        td = getdate(r["training_date"]) if r.get("training_date") else None
        vd = getdate(r["valid_to"]) if r.get("valid_to") else None

        record = {
            "name": r["name"],
            "training_date": td,
            "valid_to": vd,
            "docstatus": r.get("docstatus"),
        }

        if r.get("docstatus") == 1 and idx[key]["submitted"] is None:
            idx[key]["submitted"] = record
            continue

        if r.get("docstatus") != 1 and td and td >= today and idx[key]["scheduled"] is None:
            idx[key]["scheduled"] = record

    return idx


def _build_cell_payload(emp, induction, idx, today, warn_date):
    entry = idx.get((emp, induction)) or {
        "submitted": None,
        "scheduled": None,
    }

    sub = entry.get("submitted")
    sch = entry.get("scheduled")

    payload = {
        "status": "red",
        "expiry": None,
        "days": None,
        "last": None,
        "scheduled": None,
        "submitted_record": None,
        "scheduled_record": None,
    }

    if sch and sch.get("training_date"):
        payload["scheduled"] = sch["training_date"].isoformat()
        payload["scheduled_record"] = sch.get("name")

    if not sub:
        return payload

    payload["submitted_record"] = sub.get("name")

    if sub.get("training_date"):
        payload["last"] = sub["training_date"].isoformat()

    if not sub.get("valid_to"):
        payload["status"] = "red"
        return payload

    expiry = sub["valid_to"]
    payload["expiry"] = expiry.isoformat()
    payload["days"] = (expiry - today).days

    if expiry < today:
        payload["status"] = "red"
    elif expiry <= warn_date:
        payload["status"] = "yellow"
    else:
        payload["status"] = "green"

    return payload