# Copyright (c) 2026, BuFf0k and contributors
# For license information, please see license.txt

from __future__ import unicode_literals

from collections import OrderedDict, defaultdict
from io import BytesIO

import frappe
from frappe import _
from frappe.model.document import Document
from frappe.utils import getdate

from ir.industrial_relations.utils import autoname_planning_document


class SiteOrganogram(Document):
    def autoname(self):
        autoname_planning_document(self, "location")

    def validate(self):
        self.validate_effective_dates()
        normalize_group_structure(self)
        normalize_mappings(self)
        normalize_reporting_lines(self)

    def before_submit(self):
        normalize_group_structure(self)
        normalize_mappings(self)
        normalize_reporting_lines(self)

    def validate_effective_dates(self):
        if (
            self.effective_from
            and self.effective_until
            and getdate(self.effective_until) < getdate(self.effective_from)
        ):
            frappe.throw(_("Effective Until cannot be before Effective From."))


# -------------------------------------------------------------------
# Generic helpers
# -------------------------------------------------------------------

def _as_list(value):
    if value is None:
        return []
    if isinstance(value, list):
        return value
    if isinstance(value, str):
        try:
            parsed = frappe.parse_json(value)
            return parsed if isinstance(parsed, list) else []
        except Exception:
            return []
    return []


def _clean(value):
    if value is None:
        return ""
    return str(value).strip()


def _safe_int(value, default=0):
    try:
        return int(value or default)
    except Exception:
        return default


def _parse_row_key(row_key):
    row_key = _clean(row_key)

    if row_key.startswith("ASSET::"):
        return {
            "kind": "Asset",
            "asset": row_key[len("ASSET::"):],
            "designation": "",
        }

    if row_key.startswith("DESIG::"):
        rest = row_key[len("DESIG::"):]
        parts = rest.split("::")
        return {
            "kind": "Designation",
            "asset": "",
            "designation": parts[0] or "",
            "token": parts[1] if len(parts) > 1 else "",
        }

    return {
        "kind": "Unknown",
        "asset": "",
        "designation": "",
    }


def _row_key_for_asset(asset):
    return f"ASSET::{_clean(asset)}"


def _row_key_for_designation(label, token=None):
    label = _clean(label) or "Unlinked Role"
    token = _clean(token) or frappe.generate_hash(length=6)
    return f"DESIG::{label}::{token}"


def _derive_row_key(row):
    existing = _clean(getattr(row, "row_key", None))
    if existing:
        return existing

    row_type = _clean(getattr(row, "row_type", None))
    asset = _clean(getattr(row, "asset", None))
    label = _clean(getattr(row, "row_label", None))

    if row_type == "Asset" or asset:
        if asset:
            return _row_key_for_asset(asset)
        return f"MISSING_ASSET::{frappe.generate_hash(length=6)}"

    return _row_key_for_designation(label or "Unlinked Role")


def _asset_display(asset_id):
    if not asset_id:
        return ""

    values = frappe.db.get_value(
        "Asset",
        asset_id,
        ["name", "item_name", "asset_category"],
        as_dict=True,
    )

    if not values:
        return asset_id

    parts = [values.name]
    if values.item_name:
        parts.append(values.item_name)
    elif values.asset_category:
        parts.append(values.asset_category)

    return " — ".join(parts)


def _employee_exists(employee_id):
    if not employee_id:
        return False
    return bool(frappe.db.exists("Employee", employee_id))


def _asset_exists(asset_id):
    if not asset_id:
        return False
    return bool(frappe.db.exists("Asset", asset_id))


def _new_group_key():
    return f"GRP::{frappe.generate_hash(length=10)}"


def normalize_group_structure(doc):
    """Assign stable keys to headings and mirror them onto mapping rows."""
    headings = getattr(doc, "group_headings", None) or []
    mappings = getattr(doc, "shift_mappings", None) or []

    by_key = {}
    by_label = {}

    for heading in headings:
        label = _clean(getattr(heading, "group", None))
        key = _clean(getattr(heading, "group_key", None))

        if not key:
            key = _new_group_key()
            heading.group_key = key

        if key in by_key and by_key[key] is not heading:
            key = _new_group_key()
            heading.group_key = key

        by_key[key] = heading
        if label and label not in by_label:
            by_label[label] = heading

    for row in mappings:
        key = _clean(getattr(row, "group_key", None))
        label = _clean(getattr(row, "group", None))
        heading = by_key.get(key) if key else None

        if not heading and label:
            heading = by_label.get(label)

        if heading:
            row.group_key = heading.group_key
            row.group = heading.group


def normalize_reporting_lines(doc):
    """Repair reporting-line endpoints without deleting user data."""
    headings = getattr(doc, "group_headings", None) or []
    lines = getattr(doc, "reporting_lines", None) or []

    by_key = {
        _clean(getattr(row, "group_key", None)): row
        for row in headings
        if _clean(getattr(row, "group_key", None))
    }
    by_label = {}
    for row in headings:
        label = _clean(getattr(row, "group", None))
        if label and label not in by_label:
            by_label[label] = row

    for index, line in enumerate(lines, start=1):
        for prefix in ("source", "target"):
            key_field = f"{prefix}_group_key"
            label_field = f"{prefix}_group"
            scope_field = f"{prefix}_scope"
            shift_field = f"{prefix}_shift"

            key = _clean(getattr(line, key_field, None))
            label = _clean(getattr(line, label_field, None))
            heading = by_key.get(key) if key else None

            if not heading and label:
                heading = by_label.get(label)

            if heading:
                setattr(line, key_field, heading.group_key)
                setattr(line, label_field, heading.group)

            scope = _clean(getattr(line, scope_field, None)) or "Heading"
            if scope not in ("Heading", "Shift"):
                scope = "Heading"
            setattr(line, scope_field, scope)

            if scope == "Heading":
                setattr(line, shift_field, "")

        line.line_type = _clean(getattr(line, "line_type", None)) or "Solid"
        line.source_anchor = _clean(getattr(line, "source_anchor", None)) or "Auto"
        line.target_anchor = _clean(getattr(line, "target_anchor", None)) or "Auto"
        if not _safe_int(getattr(line, "line_order", 0), 0):
            line.line_order = index


def normalize_mappings(doc):
    """
    Server-side safety net.

    Important:
    - Never delete mapping rows here.
    - Preserve organogram structure.
    - Ensure row_key, row_type, row_label, missing flags, and row_order are sane.
    """

    rows = getattr(doc, "shift_mappings", None) or []
    if not rows:
        return

    # First pass: repair row identity and labels.
    for row in rows:
        row.row_key = _derive_row_key(row)

        info = _parse_row_key(row.row_key)
        current_type = _clean(getattr(row, "row_type", None))

        if info["kind"] == "Asset":
            row.row_type = "Asset"

            asset_from_key = info.get("asset") or ""
            if not _clean(getattr(row, "asset", None)):
                row.asset = asset_from_key if _asset_exists(asset_from_key) else ""

            if not _clean(getattr(row, "row_label", None)):
                row.row_label = _asset_display(asset_from_key) or asset_from_key or "Missing"

            if _clean(getattr(row, "asset", None)) and _asset_exists(row.asset):
                row.missing_asset = 0
            else:
                row.missing_asset = 1

        elif info["kind"] == "Designation":
            row.row_type = "Designation"
            row.asset = ""
            row.missing_asset = 0
            row.designation = ""

            if not _clean(getattr(row, "row_label", None)):
                row.row_label = info.get("designation") or "Unlinked Role"

        else:
            if current_type == "Asset":
                row.row_type = "Asset"
                row.missing_asset = 1
                if not _clean(getattr(row, "row_label", None)):
                    row.row_label = "Missing"
            else:
                row.row_type = "Designation"
                row.asset = ""
                row.missing_asset = 0
                row.designation = ""
                if not _clean(getattr(row, "row_label", None)):
                    row.row_label = "Unlinked Role"

        # A Designation can never be Spare/Swing - only a physical Asset can be.
        if row.row_type != "Asset":
            row.spare_swing = 0
        else:
            row.spare_swing = _safe_int(getattr(row, "spare_swing", 0), 0)

        if _safe_int(getattr(row, "spare_swing", 0), 0):
            row.employee = ""
            row.missing_employee = 0

        if _clean(getattr(row, "employee", None)) and _employee_exists(row.employee):
            row.missing_employee = 0
        elif _clean(getattr(row, "employee", None)):
            row.employee = ""
            row.missing_employee = 1
        else:
            row.missing_employee = _safe_int(getattr(row, "missing_employee", 0), 0)

    # Spare/Swing is a property of the physical row (one row_key spans one
    # Site Organogram Mappings record per active shift) - if any shift-row for
    # a given row_key was marked Spare/Swing, every sibling shift-row for that
    # same row_key must be too, and any employee left on them must be cleared.
    rows_by_key = defaultdict(list)
    for row in rows:
        row_key = _clean(getattr(row, "row_key", None))
        if row_key:
            rows_by_key[row_key].append(row)

    for row_key, key_rows in rows_by_key.items():
        if any(_safe_int(getattr(r, "spare_swing", 0), 0) for r in key_rows):
            for row in key_rows:
                row.spare_swing = 1
                row.employee = ""
                row.missing_employee = 0

    # Default Designation is likewise a property of the physical row, not of
    # any one shift-slot - if any sibling shift-row for a given Asset row_key
    # has a Designation set, every sibling shift-row for that same row_key
    # must carry the same one.
    for row_key, key_rows in rows_by_key.items():
        designations = {
            _clean(getattr(r, "designation", None))
            for r in key_rows
            if _clean(getattr(r, "row_type", None)) == "Asset" and _clean(getattr(r, "designation", None))
        }
        if designations:
            designation = sorted(designations)[0]
            for row in key_rows:
                if _clean(getattr(row, "row_type", None)) == "Asset":
                    row.designation = designation

    # Second pass: stable row order per group.
    groups = defaultdict(list)

    for row in rows:
        group = _clean(getattr(row, "group", None))
        row_key = _clean(getattr(row, "row_key", None))
        if not group or not row_key:
            continue
        groups[group].append(row)

    for group, group_rows in groups.items():
        key_order = OrderedDict()

        for row in sorted(
            group_rows,
            key=lambda r: (
                _safe_int(getattr(r, "row_order", 0), 999999) or 999999,
                _safe_int(getattr(r, "idx", 0), 999999) or 999999,
            ),
        ):
            if row.row_key not in key_order:
                key_order[row.row_key] = len(key_order) + 1

        for row in group_rows:
            row.row_order = key_order.get(row.row_key) or 1


# -------------------------------------------------------------------
# Branch / Location helpers
# -------------------------------------------------------------------

@frappe.whitelist()
def get_matching_location_for_branch(branch):
    if not branch:
        return None
    return branch if frappe.db.exists("Location", branch) else None


# -------------------------------------------------------------------
# Employee / Asset sync
# -------------------------------------------------------------------

@frappe.whitelist()
def sync_employees(branch, current_employees=None, auto_employees=None):
    current_employees = _as_list(current_employees)
    auto_employees = _as_list(auto_employees)

    if not branch:
        return {
            "to_add": [],
            "to_remove": sorted(list(set(auto_employees))),
        }

    rows = frappe.get_all(
        "Employee",
        filters={
            "status": "Active",
            "branch": branch,
        },
        fields=["name", "employee_name", "designation"],
        order_by="employee_name asc",
    )

    should_auto = {r.name: r for r in rows}
    current_set = set([x for x in current_employees if x])
    auto_set = set([x for x in auto_employees if x])

    to_remove = sorted(list(auto_set - set(should_auto.keys())))

    to_add = []
    for emp_id, row in should_auto.items():
        if emp_id not in current_set:
            to_add.append(
                {
                    "employee": emp_id,
                    "employee_name": row.employee_name or "",
                    "designation": row.designation or "",
                    # Always == branch here (that's the filter above) - kept
                    # so an auto-added row has the same shape as a manually-
                    # added one (see get_employee_details()), rather than
                    # the Designer needing to special-case which rows have it.
                    "branch": branch,
                }
            )

    return {
        "to_add": to_add,
        "to_remove": to_remove,
    }


@frappe.whitelist()
def get_employee_details(employee):
    if not employee:
        return {}

    doc = frappe.get_doc("Employee", employee)

    return {
        "employee_name": doc.employee_name,
        "designation": doc.designation,
        "branch": doc.branch,
    }


@frappe.whitelist()
def sync_assets(location, asset_categories=None, current_assets=None, auto_assets=None):
    asset_categories = _as_list(asset_categories)
    current_assets = _as_list(current_assets)
    auto_assets = _as_list(auto_assets)

    auto_set = set([x for x in auto_assets if x])

    if not location:
        return {
            "to_add": [],
            "to_remove": sorted(list(auto_set)),
        }

    filters = {
        "docstatus": 1,
        "location": location,
    }

    if asset_categories:
        filters["asset_category"] = ["in", list({c for c in asset_categories if c})]

    rows = frappe.get_all(
        "Asset",
        filters=filters,
        fields=["name", "item_name", "asset_category"],
        order_by="item_name asc",
    )

    should_auto = {r.name: r for r in rows}
    current_set = set([x for x in current_assets if x])

    to_remove = sorted(list(auto_set - set(should_auto.keys())))

    to_add = []
    for asset_id, row in should_auto.items():
        if asset_id not in current_set:
            to_add.append(
                {
                    "asset": asset_id,
                    "item_name": row.item_name or "",
                    "asset_category": row.asset_category or "",
                }
            )

    return {
        "to_add": to_add,
        "to_remove": to_remove,
    }


@frappe.whitelist()
def get_asset_details(asset):
    if not asset:
        return {}

    doc = frappe.get_doc("Asset", asset)

    return {
        "item_name": doc.item_name,
        "asset_category": doc.asset_category,
    }


@frappe.whitelist()
def debug_assets_query(location, asset_categories=None):
    cats = _as_list(asset_categories)

    filters = {
        "docstatus": 1,
    }

    if location:
        filters["location"] = location

    if cats:
        filters["asset_category"] = ["in", list({c for c in cats if c})]

    count = frappe.db.count("Asset", filters)

    sample = frappe.get_all(
        "Asset",
        filters=filters,
        fields=["name", "item_name", "asset_category", "location", "docstatus"],
        limit_page_length=10,
        order_by="modified desc",
    )

    return {
        "location_received": location,
        "categories_received": cats,
        "filters_used": filters,
        "count": count,
        "sample": sample,
    }


# -------------------------------------------------------------------
# Template / Clone helpers
# -------------------------------------------------------------------

@frappe.whitelist()
def list_recent_site_organograms_for_branch(branch, exclude_name=None, limit=5):
    if not branch:
        return []

    try:
        limit = int(limit or 5)
    except Exception:
        limit = 5

    limit = max(1, min(20, limit))

    filters = {
        "branch": branch,
    }

    if exclude_name:
        filters["name"] = ["!=", exclude_name]

    rows = frappe.get_all(
        "Site Organogram",
        filters=filters,
        fields=["name", "modified"],
        order_by="modified desc",
        limit_page_length=limit,
    )

    return rows or []


@frappe.whitelist()
def get_site_organogram_template(source_name):
    """
    Return a safe clone payload.

    Important:
    - Copies the structure and mapping table exactly.
    - Does not decide what is valid for the target document.
    - The target JS reconciles against the current branch/location pools without deleting rows.
    """

    if not source_name:
        return {}

    doc = frappe.get_doc("Site Organogram", source_name)
    normalize_group_structure(doc)
    normalize_mappings(doc)
    normalize_reporting_lines(doc)

    return {
        "location": getattr(doc, "location", None),
        "group_headings": [
            {
                "group_key": getattr(r, "group_key", None),
                "group": r.group,
                "shift_design": getattr(r, "shift_design", None),
            }
            for r in (getattr(doc, "group_headings", None) or [])
        ],
        "asset_categories": [
            {
                "asset_cateogories": r.asset_cateogories,
            }
            for r in (getattr(doc, "asset_categories", None) or [])
        ],
        "employees": [
            {
                "employee": r.employee,
                "employee_name": r.employee_name,
                "designation": r.designation,
            }
            for r in (getattr(doc, "employees", None) or [])
        ],
        "assets": [
            {
                "asset": r.asset,
                "item_name": r.item_name,
                "asset_category": r.asset_category,
            }
            for r in (getattr(doc, "assets", None) or [])
        ],
        "shift_mappings": [
            {
                "group_key": getattr(r, "group_key", None),
                "group": r.group,
                "shift": r.shift,
                "employee": r.employee,
                "asset": r.asset,
                "designation": getattr(r, "designation", None),
                "row_key": getattr(r, "row_key", None),
                "row_order": getattr(r, "row_order", None),
                "row_label": getattr(r, "row_label", None),
                "row_type": getattr(r, "row_type", None),
                "spare_swing": getattr(r, "spare_swing", 0),
                "missing_asset": getattr(r, "missing_asset", 0),
                "missing_employee": getattr(r, "missing_employee", 0),
            }
            for r in (getattr(doc, "shift_mappings", None) or [])
        ],
        "reporting_lines": [
            {
                "source_group_key": getattr(r, "source_group_key", None),
                "source_group": getattr(r, "source_group", None),
                "source_scope": getattr(r, "source_scope", None),
                "source_shift": getattr(r, "source_shift", None),
                "target_group_key": getattr(r, "target_group_key", None),
                "target_group": getattr(r, "target_group", None),
                "target_scope": getattr(r, "target_scope", None),
                "target_shift": getattr(r, "target_shift", None),
                "line_type": getattr(r, "line_type", None),
                "label": getattr(r, "label", None),
                "source_anchor": getattr(r, "source_anchor", None),
                "target_anchor": getattr(r, "target_anchor", None),
                "line_order": getattr(r, "line_order", None),
            }
            for r in (getattr(doc, "reporting_lines", None) or [])
        ],
    }


@frappe.whitelist()
def get_site_plan_template(site_plan_name):
    """Return a Site Plan's structure (Groups, Slots, Reporting Lines) shaped for
    the Organogram Designer to merge into its own state.

    Unlike get_site_organogram_template (which clones another *Organogram*,
    including its real Employee/Asset assignments), a Site Plan carries no
    Employees or Assets by design - only the shape (Group Headings and
    Reporting Lines) and Slots (what should exist: a Designation, or an Asset
    of a given Category). Slots are expanded here into one blank/vacant
    mapping row per shift column of their group, mirroring add_row()'s own
    per-shift expansion in the JS, so the caller can merge them straight into
    shift_mappings without any further shift-expansion logic of its own.
    """
    if not site_plan_name:
        return {}

    plan = frappe.get_doc("Site Plan", site_plan_name)
    groups = plan.groups or []
    group_names = {row.group_key: row.group for row in groups}
    shift_counts = {row.group_key: _shift_design_team_count(row.shift_design) for row in groups}

    asset_categories = sorted({
        _clean(slot.asset_category)
        for slot in (plan.slots or [])
        if slot.row_type == "Asset" and _clean(slot.asset_category)
    })

    shift_mappings = []
    for slot in (plan.slots or []):
        count = max(0, min(20, shift_counts.get(slot.group_key, 0)))

        # normalize_mappings() (run on every save) parses row_key to decide a
        # row's kind and *overwrites* row_type/designation/row_label from
        # that parse - it only recognises its own ASSET::/DESIG:: formats, so
        # a Plan Slot's own SLOT::<hash> key must be translated into one of
        # those here, not passed through as-is, or the row loses its
        # designation/label on the very first save. The hash suffix of the
        # Slot's own row_key is reused as the token, so this stays stable
        # (and idempotent for the caller's own re-populate dedup) across
        # repeated calls for the same Slot.
        token = (slot.row_key or "").split("::")[-1] or frappe.generate_hash(length=6)

        if slot.row_type == "Designation":
            label = slot.designation or slot.row_label or "Unlinked Role"
            row_key = _row_key_for_designation(label, token=token)
            row_label = label
        else:
            # No real Asset exists yet, so there is no ASSET::<id> to form -
            # this lands in normalize_mappings()'s existing "unresolved Asset
            # row" fallback (missing_asset=1) the same way a row whose linked
            # Asset was since deleted would, until a real Asset is assigned.
            row_key = f"MISSING_ASSET::{token}"
            row_label = slot.row_label or slot.asset_category or "Missing"

        for shift in (f"Shift {x}" for x in SHIFT_LETTERS[:count]):
            shift_mappings.append(
                {
                    "group_key": slot.group_key,
                    "group": group_names.get(slot.group_key, ""),
                    "shift": shift,
                    "employee": "",
                    "asset": "",
                    "designation": slot.designation if slot.row_type == "Asset" else "",
                    "row_key": row_key,
                    "row_order": slot.row_order,
                    "row_label": row_label,
                    "row_type": slot.row_type,
                    "spare_swing": slot.spare_swing,
                    "missing_asset": 0,
                    "missing_employee": 0,
                }
            )

    return {
        "plan_name": plan.plan_name,
        "branch": plan.branch,
        "location": plan.location,
        "effective_from": plan.effective_from,
        "effective_until": plan.effective_until,
        "asset_categories": asset_categories,
        "group_headings": [
            {
                "group_key": row.group_key,
                "group": row.group,
                "shift_design": row.shift_design,
            }
            for row in groups
        ],
        "shift_mappings": shift_mappings,
        "reporting_lines": [
            {
                "source_group_key": row.source_group_key,
                "source_group": row.source_group,
                "source_scope": row.source_scope,
                "source_shift": row.source_shift,
                "target_group_key": row.target_group_key,
                "target_group": row.target_group,
                "target_scope": row.target_scope,
                "target_shift": row.target_shift,
                "line_type": row.line_type,
                "label": row.label,
                "source_anchor": row.source_anchor,
                "target_anchor": row.target_anchor,
                "line_order": row.line_order,
            }
            for row in (plan.reporting_lines or [])
        ],
    }


# -------------------------------------------------------------------
# Excel export
# -------------------------------------------------------------------

SHIFT_LETTERS = [chr(65 + i) for i in range(20)]


def _shift_design_team_count(shift_design):
    if not shift_design:
        return 0
    return _safe_int(frappe.db.get_value("Shift Design", shift_design, "number_of_teams"), 0)


def _group_shift_labels(doc, group_row):
    count = max(0, min(20, _shift_design_team_count(getattr(group_row, "shift_design", None))))
    return [f"Shift {x}" for x in SHIFT_LETTERS[:count]]


def _split_employee_name(full_name):
    full_name = _clean(full_name)

    if not full_name:
        return "", ""

    parts = full_name.split()

    if len(parts) == 1:
        return parts[0], ""

    return " ".join(parts[:-1]), parts[-1]


def _employee_lookup(doc):
    lookup = {}

    for row in getattr(doc, "employees", None) or []:
        if not row.employee:
            continue

        first_names, surname = _split_employee_name(row.employee_name or row.employee)

        lookup[row.employee] = {
            "employee": row.employee,
            "employee_name": row.employee_name or row.employee,
            "first_names": first_names,
            "surname": surname,
            "designation": row.designation or "",
        }

    return lookup


def _asset_lookup(doc):
    lookup = {}

    for row in getattr(doc, "assets", None) or []:
        if not row.asset:
            continue

        lookup[row.asset] = {
            "asset": row.asset,
            "item_name": row.item_name or "",
            "asset_category": row.asset_category or "",
        }

    return lookup


def _get_admin_exceptions(doc):
    """Employees/Assets allocated somewhere in this organogram whose own
    Employee.branch / Asset.location record disagrees with this organogram's
    branch/location - i.e. their admin record says they belong elsewhere.
    Returns (employee_rows, asset_rows), each a list of plain lists matching
    the export headers below.
    """
    rows = getattr(doc, "shift_mappings", None) or []

    employee_ids = sorted({row.employee for row in rows if row.employee})
    asset_ids = sorted({
        row.asset for row in rows
        if _clean(getattr(row, "row_type", None)) == "Asset" and row.asset
    })

    employee_rows = []
    if employee_ids:
        for emp in frappe.get_all(
            "Employee",
            filters={"name": ["in", employee_ids]},
            fields=["name", "employee_name", "designation", "branch"],
        ):
            if (emp.branch or "") != (doc.branch or ""):
                employee_rows.append([emp.name, emp.employee_name or "", emp.designation or "", emp.branch or ""])

    asset_rows = []
    if asset_ids:
        for asset in frappe.get_all(
            "Asset",
            filters={"name": ["in", asset_ids]},
            fields=["name", "item_name", "asset_category", "location"],
        ):
            if (asset.location or "") != (doc.location or ""):
                asset_rows.append([asset.name, asset.item_name or "", asset.asset_category or "", asset.location or ""])

    return employee_rows, asset_rows


def _get_vacancy_summary(doc):
    """Single source of truth for vacancy counting, at shift-slot granularity
    (each empty shift-cell is one vacancy). A Designation row with no employee
    is always vacant; an Asset row with no employee is vacant unless it's
    marked Spare/Swing (a spare asset doesn't need staffing). An Asset row
    with a Designation set counts toward that Designation too - it's still
    listed in vacant_assets (which specific asset), on top of by_designation
    (how many of that role overall) - the two answer different questions, so
    a vacant Asset can appear in both, but `total` counts each vacant
    shift-cell exactly once regardless of how many buckets it lands in.
    Returns: {"by_designation": {label: count}, "vacant_assets": [rows...], "total": n}
    """
    rows = getattr(doc, "shift_mappings", None) or []
    assets = _asset_lookup(doc)

    by_designation = defaultdict(int)
    vacant_assets = []
    total = 0

    for row in rows:
        if row.employee:
            continue

        row_type = _clean(getattr(row, "row_type", None))

        if row_type == "Designation":
            info = _parse_row_key(getattr(row, "row_key", None))
            label = info.get("designation") or _clean(getattr(row, "row_label", None)) or "Unlinked Role"
            by_designation[label] += 1
            total += 1

        elif row_type == "Asset" and not _safe_int(getattr(row, "spare_swing", 0), 0):
            asset = assets.get(row.asset, {})
            vacant_assets.append([
                row.asset or "",
                asset.get("item_name") or "",
                asset.get("asset_category") or "",
                row.shift or "",
            ])

            designation = _clean(getattr(row, "designation", None))
            if designation:
                by_designation[designation] += 1

            total += 1

    return {
        "by_designation": dict(by_designation),
        "vacant_assets": vacant_assets,
        "total": total,
    }


def _iter_designation_slots(doc):
    """Yield one dict per staffable FTE slot - a Designation row, or an
    Asset row with a default Designation set (Spare/Swing Asset rows are
    excluded, same as _get_vacancy_summary()'s semantics: a spare asset
    isn't a staffing requirement). Asset rows with no Designation set are
    skipped too - there's nothing to attribute them to.

    Single shared row-classification used by get_designation_headcounts(),
    get_designation_mismatches() and get_designation_slots_by_group(), so
    "what counts as this row's Designation" is defined in exactly one place.
    Yields: {group_key, group, shift, row_label, employee, designation}
    """
    for row in getattr(doc, "shift_mappings", None) or []:
        row_type = _clean(getattr(row, "row_type", None))

        if row_type == "Designation":
            info = _parse_row_key(getattr(row, "row_key", None))
            designation = info.get("designation") or _clean(getattr(row, "row_label", None)) or "Unlinked Role"
        elif row_type == "Asset":
            if _safe_int(getattr(row, "spare_swing", 0), 0):
                continue
            designation = _clean(getattr(row, "designation", None))
            if not designation:
                continue
        else:
            continue

        yield {
            "group_key": getattr(row, "group_key", None) or "",
            "group": row.group or "",
            "shift": row.shift or "",
            "row_label": row.row_label or "",
            "employee": row.employee or "",
            "designation": designation,
        }


def get_designation_headcounts(doc):
    """Per-Designation FTE headcount (filled/vacant/total) across the whole
    Organogram - drives Site Budget's Designation costing table.
    Returns: {designation: {"filled": n, "vacant": n, "total": n}}
    """
    counts = defaultdict(lambda: {"filled": 0, "vacant": 0, "total": 0})

    for slot in _iter_designation_slots(doc):
        bucket = counts[slot["designation"]]
        bucket["total"] += 1
        if slot["employee"]:
            bucket["filled"] += 1
        else:
            bucket["vacant"] += 1

    return {label: dict(values) for label, values in counts.items()}


def get_designation_mismatches(doc):
    """Filled slots where the assigned Employee's actual Designation doesn't
    match the role's expected Designation - e.g. a Multi-Skilled Operator
    assigned against a Dozer whose Designation is set to Dozer Operator.
    """
    employees = _employee_lookup(doc)
    mismatches = []

    for slot in _iter_designation_slots(doc):
        if not slot["employee"]:
            continue

        employee_info = employees.get(slot["employee"], {})
        actual = _clean(employee_info.get("designation"))

        if actual and actual != slot["designation"]:
            mismatches.append({
                "group": slot["group"],
                "shift": slot["shift"],
                "row_label": slot["row_label"],
                "employee": slot["employee"],
                "employee_name": employee_info.get("employee_name") or slot["employee"],
                "expected_designation": slot["designation"],
                "actual_designation": actual,
            })

    return mismatches


def get_employee_branch_exceptions(doc):
    """Public wrapper over _get_admin_exceptions()'s employee half - Employees
    assigned somewhere in this Organogram whose own Employee.branch disagrees
    with this Organogram's branch (e.g. a shared Engineering/Maintenance
    resource who genuinely works more than one site). Already surfaced on
    the Organogram Designer's own report panel and Excel export as "Admin
    Exceptions - Employees"; Site Budget's summaries reuse this same
    detection rather than re-implementing it, so the two can't disagree.
    Returns a list of dicts (not the plain [name, ...] lists
    _get_admin_exceptions() itself returns, which are shaped for that
    Excel sheet specifically).
    """
    employee_rows, _asset_rows = _get_admin_exceptions(doc)
    return [
        {
            "employee": row[0],
            "employee_name": row[1],
            "designation": row[2],
            "branch": row[3],
        }
        for row in employee_rows
    ]


def get_designation_slots_by_group(doc):
    """Every staffable slot, grouped by heading (group_key) with that
    heading's own Shift Design - Site Budget uses this to cost each
    Designation per-heading, since different headings can carry different
    Shift Designs with different pay periods.

    Also broken down by `shift` ("Shift A"/"Shift B"/...) within each group,
    since Site Budget's overtime estimate needs to know which specific team
    within that heading's Shift Design each slot belongs to (team order ==
    shift-letter order, same as everywhere else in the app) - a slot's
    overtime hours differ by team, not just by heading.

    Returns: {group_key: {"shift_design": x, "designation_counts": {designation: count},
    "by_shift": {shift: {designation: count}}}}
    """
    group_shift_design = {
        row.group_key: row.shift_design
        for row in getattr(doc, "group_headings", None) or []
    }

    by_group = defaultdict(
        lambda: {
            "shift_design": "",
            "designation_counts": defaultdict(int),
            "by_shift": defaultdict(lambda: defaultdict(int)),
        }
    )

    for slot in _iter_designation_slots(doc):
        bucket = by_group[slot["group_key"]]
        bucket["shift_design"] = group_shift_design.get(slot["group_key"], "")
        bucket["designation_counts"][slot["designation"]] += 1
        bucket["by_shift"][slot["shift"]][slot["designation"]] += 1

    return {
        group_key: {
            "shift_design": value["shift_design"],
            "designation_counts": dict(value["designation_counts"]),
            "by_shift": {
                shift: dict(designation_counts) for shift, designation_counts in value["by_shift"].items()
            },
        }
        for group_key, value in by_group.items()
    }


def _mapping_indexes(doc):
    by_slot = {}
    row_keys_by_group = defaultdict(OrderedDict)

    rows = sorted(
        getattr(doc, "shift_mappings", None) or [],
        key=lambda row: (
            _clean(getattr(row, "group", None)),
            _safe_int(getattr(row, "row_order", 0), 999999) or 999999,
            _safe_int(getattr(row, "idx", 0), 999999) or 999999,
        ),
    )

    for row in rows:
        group = _clean(getattr(row, "group", None))
        shift = _clean(getattr(row, "shift", None))
        row_key = _clean(getattr(row, "row_key", None))

        if not group or not shift or not row_key:
            continue

        row_keys_by_group[group].setdefault(row_key, row)
        by_slot[(group, shift, row_key)] = row

    return by_slot, row_keys_by_group


def _employee_display(emp):
    if not emp:
        return "", ""
    return emp.get("employee_name") or emp.get("employee") or "", emp.get("employee") or ""


def _style_range_border(ws, min_row, max_row, min_col, max_col, border):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border


def _row_label_for_export(row, assets):
    row_type = _clean(getattr(row, "row_type", None))
    row_label = _clean(getattr(row, "row_label", None))
    missing_asset = _safe_int(getattr(row, "missing_asset", 0), 0)

    if row_type == "Asset":
        if missing_asset:
            return "Missing", row_label or "Missing"

        asset_id = _clean(getattr(row, "asset", None))
        asset = assets.get(asset_id)

        if asset:
            return asset.get("asset") or asset_id, asset.get("item_name") or asset.get("asset_category") or ""

        return asset_id or "Missing", row_label or ""

    return row_label or "Designation", ""


def _write_simple_list(ws, row_no, title, headers, rows, styles):
    total_cols = max(1, len(headers))

    ws.merge_cells(start_row=row_no, start_column=1, end_row=row_no, end_column=total_cols)
    c = ws.cell(row_no, 1, title)
    c.font = styles["section_font"]
    c.alignment = styles["center"]
    c.fill = styles["section_fill"]
    row_no += 1

    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row_no, idx, header)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["center"]

    row_no += 1
    data_start = row_no

    if rows:
        for item in rows:
            for idx, value in enumerate(item, start=1):
                ws.cell(row_no, idx, value)
                ws.cell(row_no, idx).alignment = styles["wrap"]
            row_no += 1
    else:
        ws.cell(row_no, 1, "None")
        row_no += 1

    _style_range_border(ws, data_start - 2, row_no - 1, 1, total_cols, styles["thin_border"])

    return row_no + 1


def _write_group(ws, row_no, doc, group_row, shifts, row_keys, by_slot, employees, assets, styles):
    group = group_row.group
    total_cols = max(4, 2 + (len(shifts) * 2))

    ws.merge_cells(start_row=row_no, start_column=1, end_row=row_no, end_column=total_cols)
    cell = ws.cell(row_no, 1, group.upper())
    cell.font = styles["section_font"]
    cell.alignment = styles["center"]
    cell.fill = styles["section_fill"]
    row_no += 1

    ws.cell(row_no, 1, "")
    ws.cell(row_no, 2, "")

    col = 3
    for shift in shifts:
        ws.merge_cells(start_row=row_no, start_column=col, end_row=row_no, end_column=col + 1)
        c = ws.cell(row_no, col, shift.upper())
        c.font = styles["shift_font"]
        c.alignment = styles["center"]
        c.fill = styles["shift_fill"]
        col += 2

    row_no += 1

    ws.cell(row_no, 1, "ASSET / DESIGNATION")
    ws.cell(row_no, 2, "DESCRIPTION")
    col = 3

    for _shift in shifts:
        ws.cell(row_no, col, "OPERATOR NAME")
        ws.cell(row_no, col + 1, "COY NO")
        col += 2

    for c in range(1, total_cols + 1):
        ws.cell(row_no, c).font = styles["header_font"]
        ws.cell(row_no, c).fill = styles["header_fill"]
        ws.cell(row_no, c).alignment = styles["center"]

    row_no += 1
    data_start = row_no

    for row_key, row_identity in row_keys.items():
        label, desc = _row_label_for_export(row_identity, assets)
        ws.cell(row_no, 1, label)
        ws.cell(row_no, 2, desc)

        col = 3
        for shift in shifts:
            mapping = by_slot.get((group, shift, row_key))
            employee_id = mapping.employee if mapping and mapping.employee else None
            emp = employees.get(employee_id) if employee_id else None

            if mapping and _safe_int(getattr(mapping, "spare_swing", 0), 0):
                ws.cell(row_no, col, "Spare / Swing")
                ws.cell(row_no, col + 1, "")
            elif mapping and _safe_int(getattr(mapping, "missing_employee", 0), 0):
                ws.cell(row_no, col, "Missing")
                ws.cell(row_no, col + 1, "")
            else:
                name, coy_no = _employee_display(emp)
                ws.cell(row_no, col, name or "Vacant")
                ws.cell(row_no, col + 1, coy_no)

            col += 2

        row_no += 1

    data_end = max(data_start, row_no - 1)
    _style_range_border(ws, data_start - 3, data_end, 1, total_cols, styles["thin_border"])

    for r in range(data_start, row_no):
        for c in range(1, total_cols + 1):
            ws.cell(r, c).alignment = styles["wrap"]

    return row_no + 1


@frappe.whitelist()
def export_site_organogram_excel(name):
    if not name:
        frappe.throw("Site Organogram name is required.")

    doc = frappe.get_doc("Site Organogram", name)
    doc.check_permission("read")
    normalize_group_structure(doc)
    normalize_mappings(doc)
    normalize_reporting_lines(doc)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        frappe.throw("openpyxl is required for this export but is not installed.")

    wb = Workbook()
    ws = wb.active
    ws.title = (doc.branch or "Organogram")[:31]

    thin_side = Side(style="thin", color="000000")

    styles = {
        "title_font": Font(bold=True, size=14),
        "section_font": Font(bold=True, size=12),
        "shift_font": Font(bold=True, size=11),
        "header_font": Font(bold=True, size=10),
        "section_fill": PatternFill("solid", fgColor="D9EAD3"),
        "shift_fill": PatternFill("solid", fgColor="D9EAF7"),
        "header_fill": PatternFill("solid", fgColor="E7E6E6"),
        "center": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "wrap": Alignment(vertical="top", wrap_text=True),
        "thin_border": Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side),
    }

    employees = _employee_lookup(doc)
    assets = _asset_lookup(doc)
    by_slot, row_keys_by_group = _mapping_indexes(doc)

    assigned_employees = {
        row.employee
        for row in getattr(doc, "shift_mappings", None) or []
        if row.employee
    }

    assigned_assets = {
        row.asset
        for row in getattr(doc, "shift_mappings", None) or []
        if _clean(getattr(row, "row_type", None)) == "Asset"
        and row.asset
        and not _safe_int(getattr(row, "missing_asset", 0), 0)
    }

    row_no = 1
    max_shift_count = max(
        (len(_group_shift_labels(doc, row)) for row in getattr(doc, "group_headings", None) or []),
        default=0,
    )
    heading_cols = max(8, 2 + max_shift_count * 2)

    ws.merge_cells(start_row=row_no, start_column=1, end_row=row_no, end_column=heading_cols)
    ws.cell(row_no, 1, (doc.branch or doc.name or "SITE ORGANOGRAM").upper())
    ws.cell(row_no, 1).font = styles["title_font"]
    ws.cell(row_no, 1).alignment = styles["center"]
    row_no += 1

    ws.merge_cells(start_row=row_no, start_column=1, end_row=row_no, end_column=heading_cols)
    ws.cell(
        row_no,
        1,
        "Site Plan: {}  |  Location: {}  |  Effective: {} to {}".format(
            doc.site_plan or "-",
            doc.location or "-",
            doc.effective_from or "-",
            doc.effective_until or "indefinite",
        ),
    )
    ws.cell(row_no, 1).alignment = styles["center"]
    row_no += 2

    groups = [row for row in getattr(doc, "group_headings", None) or [] if row.group]

    for group_row in groups:
        group = group_row.group
        row_keys = row_keys_by_group.get(group, OrderedDict())

        if not row_keys:
            continue

        shifts = _group_shift_labels(doc, group_row)

        if not shifts:
            continue

        row_no = _write_group(ws, row_no, doc, group_row, shifts, row_keys, by_slot, employees, assets, styles)

    unallocated_employee_rows = []
    for emp_id in sorted(set(employees.keys()) - assigned_employees, key=lambda x: employees[x].get("employee_name") or x):
        emp = employees[emp_id]
        unallocated_employee_rows.append([emp["employee"], emp["first_names"], emp["surname"], emp["designation"]])

    unallocated_asset_rows = []
    for asset_id in sorted(set(assets.keys()) - assigned_assets, key=lambda x: assets[x].get("item_name") or x):
        asset = assets[asset_id]
        unallocated_asset_rows.append([asset["asset"], asset["item_name"], asset["asset_category"]])

    row_no += 1
    row_no = _write_simple_list(
        ws,
        row_no,
        "UNALLOCATED EMPLOYEES",
        ["COY NO", "NAME", "SURNAME", "DESIGNATION"],
        unallocated_employee_rows,
        styles,
    )

    row_no = _write_simple_list(
        ws,
        row_no,
        "UNALLOCATED ASSETS",
        ["PLANT NO", "MACHINE MAKE", "ASSET CATEGORY"],
        unallocated_asset_rows,
        styles,
    )

    admin_exception_employee_rows, admin_exception_asset_rows = _get_admin_exceptions(doc)

    row_no = _write_simple_list(
        ws,
        row_no,
        "ADMIN EXCEPTIONS - EMPLOYEES",
        ["COY NO", "NAME", "DESIGNATION", "EMPLOYEE'S BRANCH"],
        admin_exception_employee_rows,
        styles,
    )

    row_no = _write_simple_list(
        ws,
        row_no,
        "ADMIN EXCEPTIONS - ASSETS",
        ["PLANT NO", "MACHINE MAKE", "ASSET CATEGORY", "ASSET'S LOCATION"],
        admin_exception_asset_rows,
        styles,
    )

    vacancy_summary = _get_vacancy_summary(doc)

    vacant_designation_rows = sorted(vacancy_summary["by_designation"].items(), key=lambda item: item[0])
    row_no = _write_simple_list(
        ws,
        row_no,
        "VACANT POSITIONS PER DESIGNATION",
        ["DESIGNATION", "VACANT COUNT"],
        [[label, count] for label, count in vacant_designation_rows],
        styles,
    )

    row_no = _write_simple_list(
        ws,
        row_no,
        "VACANT ASSETS (EXCLUDING SPARE / SWING)",
        ["PLANT NO", "MACHINE MAKE", "ASSET CATEGORY", "SHIFT"],
        vacancy_summary["vacant_assets"],
        styles,
    )

    row_no = _write_simple_list(
        ws,
        row_no,
        "TOTAL UNFILLED VACANCIES",
        ["TOTAL"],
        [[vacancy_summary["total"]]],
        styles,
    )

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value is not None:
                cell.alignment = styles["wrap"]

    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        if col_idx == 1:
            ws.column_dimensions[letter].width = 22
        elif col_idx == 2:
            ws.column_dimensions[letter].width = 26
        else:
            ws.column_dimensions[letter].width = 18

    for row_idx in range(1, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 18

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    filename = f"{frappe.scrub(doc.name or 'site_organogram')}.xlsx"
    frappe.local.response.filename = filename
    frappe.local.response.filecontent = out.getvalue()
    frappe.local.response.type = "binary"
# -------------------------------------------------------------------
# Organogram Designer Page API
# -------------------------------------------------------------------

def _designer_child_rows(rows, fields):
    return [
        {field: getattr(row, field, None) for field in fields}
        for row in (rows or [])
    ]


def _designer_payload(doc):
    normalize_group_structure(doc)
    normalize_mappings(doc)
    normalize_reporting_lines(doc)

    return {
        "name": doc.name,
        "doctype": doc.doctype,
        "docstatus": doc.docstatus,
        "modified": str(doc.modified or ""),
        "branch": getattr(doc, "branch", None),
        "location": getattr(doc, "location", None),
        "site_plan": getattr(doc, "site_plan", None),
        "effective_from": str(doc.effective_from) if getattr(doc, "effective_from", None) else None,
        "effective_until": str(doc.effective_until) if getattr(doc, "effective_until", None) else None,
        "asset_categories": _designer_child_rows(
            getattr(doc, "asset_categories", None),
            ["asset_cateogories"],
        ),
        "group_headings": _designer_child_rows(
            getattr(doc, "group_headings", None),
            ["group_key", "group", "shift_design"],
        ),
        "employees": _designer_child_rows(
            getattr(doc, "employees", None),
            ["employee", "employee_name", "designation", "branch"],
        ),
        "assets": _designer_child_rows(
            getattr(doc, "assets", None),
            ["asset", "item_name", "asset_category"],
        ),
        "shift_mappings": _designer_child_rows(
            getattr(doc, "shift_mappings", None),
            [
                "group_key", "group", "shift", "employee", "asset", "designation",
                "row_key", "row_order", "row_label", "row_type", "spare_swing",
                "missing_asset", "missing_employee",
            ],
        ),
        "reporting_lines": _designer_child_rows(
            getattr(doc, "reporting_lines", None),
            [
                "source_group_key", "source_group", "source_scope", "source_shift",
                "target_group_key", "target_group", "target_scope", "target_shift",
                "line_type", "label", "source_anchor", "target_anchor", "line_order",
            ],
        ),
    }


@frappe.whitelist()
def list_site_organograms_for_designer(branch=None, limit=100):
    try:
        limit = max(1, min(int(limit or 100), 500))
    except Exception:
        limit = 100

    filters = {}
    if branch:
        filters["branch"] = branch

    return frappe.get_all(
        "Site Organogram",
        filters=filters,
        fields=["name", "branch", "location", "docstatus", "modified"],
        order_by="modified desc",
        limit_page_length=limit,
    )


@frappe.whitelist()
def get_site_organogram_designer_state(name):
    if not name:
        frappe.throw("Site Organogram name is required.")

    doc = frappe.get_doc("Site Organogram", name)
    doc.check_permission("read")
    return _designer_payload(doc)


@frappe.whitelist()
def get_site_organogram_report_summary(name):
    """Live, on-page equivalent of the Excel export's Admin Exception/Vacancy
    sections, reusing the exact same helpers so the two never drift apart."""
    if not name:
        frappe.throw("Site Organogram name is required.")

    doc = frappe.get_doc("Site Organogram", name)
    doc.check_permission("read")
    normalize_group_structure(doc)
    normalize_mappings(doc)
    normalize_reporting_lines(doc)

    employees = _employee_lookup(doc)
    assets = _asset_lookup(doc)

    assigned_employees = {row.employee for row in getattr(doc, "shift_mappings", None) or [] if row.employee}
    assigned_assets = {
        row.asset
        for row in getattr(doc, "shift_mappings", None) or []
        if _clean(getattr(row, "row_type", None)) == "Asset"
        and row.asset
        and not _safe_int(getattr(row, "missing_asset", 0), 0)
    }

    admin_exception_employees, admin_exception_assets = _get_admin_exceptions(doc)
    vacancy_summary = _get_vacancy_summary(doc)

    return {
        "unallocated_employee_count": len(set(employees.keys()) - assigned_employees),
        "unallocated_asset_count": len(set(assets.keys()) - assigned_assets),
        "admin_exception_employees": admin_exception_employees,
        "admin_exception_assets": admin_exception_assets,
        "vacant_by_designation": vacancy_summary["by_designation"],
        "vacant_assets": vacancy_summary["vacant_assets"],
        "total_unfilled_vacancies": vacancy_summary["total"],
    }


def _replace_child_table(doc, fieldname, rows, allowed_fields):
    doc.set(fieldname, [])
    for item in _as_list(rows):
        if not isinstance(item, dict):
            continue
        child = doc.append(fieldname, {})
        for field in allowed_fields:
            if field in item:
                setattr(child, field, item.get(field))


@frappe.whitelist()
def save_site_organogram_designer_state(payload):
    if isinstance(payload, str):
        payload = frappe.parse_json(payload)
    if not isinstance(payload, dict):
        frappe.throw("A valid designer payload is required.")

    name = _clean(payload.get("name"))
    expected_modified = _clean(payload.get("modified"))

    if name:
        doc = frappe.get_doc("Site Organogram", name)
        doc.check_permission("write")
        if doc.docstatus != 0:
            frappe.throw("Submitted or cancelled Site Organograms cannot be edited in the designer.")

        current_modified = str(doc.modified or "")
        if expected_modified and current_modified and expected_modified != current_modified:
            frappe.throw(
                "This Site Organogram was changed after it was loaded. Reload it before saving.",
                title="Document Changed",
            )
    else:
        doc = frappe.new_doc("Site Organogram")
        doc.check_permission("create")

    branch = _clean(payload.get("branch"))
    location = _clean(payload.get("location"))
    site_plan = _clean(payload.get("site_plan"))

    if not site_plan:
        frappe.throw("Site Plan is required. Every Site Organogram must be populated from a Site Plan.")
    if not frappe.db.exists("Site Plan", site_plan):
        frappe.throw(f"Site Plan {site_plan} does not exist.")
    if not branch:
        frappe.throw("Site is required.")
    if not location:
        frappe.throw("Location is required.")
    if not payload.get("effective_from"):
        frappe.throw("Effective From is required.")

    doc.branch = branch
    doc.location = location
    doc.site_plan = site_plan
    doc.effective_from = payload.get("effective_from") or None
    doc.effective_until = payload.get("effective_until") or None

    _replace_child_table(doc, "asset_categories", payload.get("asset_categories"), ["asset_cateogories"])
    _replace_child_table(doc, "group_headings", payload.get("group_headings"), ["group_key", "group", "shift_design"])
    _replace_child_table(doc, "employees", payload.get("employees"), ["employee", "employee_name", "designation", "branch"])
    _replace_child_table(doc, "assets", payload.get("assets"), ["asset", "item_name", "asset_category"])
    _replace_child_table(
        doc,
        "shift_mappings",
        payload.get("shift_mappings"),
        [
            "group_key", "group", "shift", "employee", "asset", "designation", "row_key",
            "row_order", "row_label", "row_type", "spare_swing",
            "missing_asset", "missing_employee",
        ],
    )
    _replace_child_table(
        doc,
        "reporting_lines",
        payload.get("reporting_lines"),
        [
            "source_group_key", "source_group", "source_scope", "source_shift",
            "target_group_key", "target_group", "target_scope", "target_shift",
            "line_type", "label", "source_anchor", "target_anchor", "line_order",
        ],
    )

    normalize_group_structure(doc)
    normalize_mappings(doc)
    normalize_reporting_lines(doc)
    doc.save()

    return _designer_payload(doc)
