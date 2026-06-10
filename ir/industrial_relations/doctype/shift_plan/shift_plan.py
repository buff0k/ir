# Copyright (c) 2026, BuFf0k and contributors
# For license information, please see license.txt

from calendar import monthrange
from collections import defaultdict
from datetime import timedelta
from io import BytesIO

import frappe
from frappe.model.document import Document
from frappe.model.naming import make_autoname
from frappe.utils import cint, flt, getdate, nowdate


class ShiftPlan(Document):
	def autoname(self):
		creation_date = getdate(nowdate())

		if self.creation:
			try:
				creation_date = getdate(self.creation)
			except Exception:
				creation_date = getdate(nowdate())

		branch_name = self.branch or "SHIFT PLAN"
		date_part = creation_date.strftime("%Y-%m-%d")

		self.name = make_autoname(f"{branch_name} - {date_part} -.###")

	def validate(self):
		self.validate_dates()
		self.set_calendar_days()
		self.set_staffing_totals()
		self.ensure_team_labels()
		self.set_rotation_defaults()

	def before_submit(self):
		if not self.output:
			self.calculate_output()

	def validate_dates(self):
		if self.calendar_start_date and self.calendar_end_date:
			if getdate(self.calendar_end_date) < getdate(self.calendar_start_date):
				frappe.throw("Calendar End Date cannot be before Calendar Start Date.")

		self.validate_day_number("Hourly Pay Period Start Day", self.hourly_pay_period_start_day)
		self.validate_day_number("Hourly Pay Period End Day", self.hourly_pay_period_end_day)
		self.validate_day_number("Salaried Pay Period Start Day", self.salaried_pay_period_start_day)
		self.validate_day_number("Salaried Pay Period End Day", self.salaried_pay_period_end_day)

	def validate_day_number(self, label, value):
		if value is None:
			return

		value = cint(value)

		if value < 1 or value > 31:
			frappe.throw(f"{label} must be between 1 and 31.")

	def set_calendar_days(self):
		if not self.calendar_start_date or not self.calendar_end_date:
			self.calendar_days = 0
			return

		start_date = getdate(self.calendar_start_date)
		end_date = getdate(self.calendar_end_date)

		self.calendar_days = (end_date - start_date).days + 1

	def set_staffing_totals(self):
		for row in self.employees:
			row.total_employees = cint(row.employees_per_shift) * cint(row.number_of_shift_teams)

			if self.meta_has_child_field("Shift Plan Staffing", "working_days"):
				if not getattr(row, "working_days", None):
					row.working_days = self.get_default_working_days_for_staffing(row)

			if self.meta_has_child_field("Shift Plan Staffing", "fixed_cost_basis"):
				if not getattr(row, "fixed_cost_basis", None):
					row.fixed_cost_basis = "Per Selected Period"

	def set_rotation_defaults(self):
		if self.parent_has_field("rotation_pattern_days") and not cint(getattr(self, "rotation_pattern_days", 0)):
			self.rotation_pattern_days = 10

		if self.parent_has_field("rotation_anchor_date") and not getattr(self, "rotation_anchor_date", None):
			self.rotation_anchor_date = self.calendar_start_date

		if self.parent_has_field("sunday_rotation_rule") and not getattr(self, "sunday_rotation_rule", None):
			self.sunday_rotation_rule = "Follow Pattern"

	def ensure_team_labels(self):
		number_of_shift_teams = cint(self.number_of_shift_teams)

		if not number_of_shift_teams:
			self.shift_team_labels = []
			return

		current_rows = list(self.shift_team_labels or [])

		if len(current_rows) > number_of_shift_teams:
			self.shift_team_labels = current_rows[:number_of_shift_teams]
			return

		for index in range(len(current_rows), number_of_shift_teams):
			row = self.append("shift_team_labels", {})
			row.team_label = f"Team {chr(65 + index)}"

	@frappe.whitelist()
	def generate_calendar(self):
		self.validate_dates()
		self.set_calendar_days()

		self.shift_calendar = []

		start_date = getdate(self.calendar_start_date)
		end_date = getdate(self.calendar_end_date)

		public_holidays = self.get_sa_public_holidays(start_date, end_date)

		current_date = start_date

		while current_date <= end_date:
			day_name = current_date.strftime("%A")
			day_type = self.get_day_type(current_date, public_holidays)
			public_holiday_name = public_holidays.get(current_date)

			planned_operating_hours, day_shift_required, night_shift_required = self.get_day_operating_values(
				day_name=day_name,
				day_type=day_type,
			)

			row = self.append("shift_calendar", {})
			row.date = current_date
			row.day_of_week = day_name
			row.day_type = day_type
			row.planned_operating_hours = planned_operating_hours
			row.day_shift_required = day_shift_required
			row.night_shift_required = night_shift_required

			if self.meta_has_child_field("Shift Plan Calendar", "public_holiday_name"):
				row.public_holiday_name = public_holiday_name or ""

			current_date = current_date + timedelta(days=1)

		return self.as_dict()

	@frappe.whitelist()
	def generate_team_rotation(self):
		if not self.shift_calendar:
			frappe.throw("Please generate the Shift Calendar before generating team rotation.")

		if not self.rotation_pattern:
			frappe.throw("Please set a Rotation Pattern before generating team rotation.")

		if not self.parent_has_field("team_rotation"):
			return self.as_dict()

		self.team_rotation = []

		rotation_map = self.expand_rotation_pattern_by_date_and_cost_group()

		for calendar_row in self.shift_calendar:
			row_date = getdate(calendar_row.date)

			for cost_group in sorted(rotation_map.get(row_date, {}).keys()):
				assignments = rotation_map.get(row_date, {}).get(cost_group, {})

				day_teams = []
				night_teams = []
				off_teams = []

				for team_label, assignment in assignments.items():
					if assignment == "Day":
						day_teams.append(team_label)
					elif assignment == "Night":
						night_teams.append(team_label)
					else:
						off_teams.append(team_label)

				row = self.append("team_rotation", {})
				row.date = calendar_row.date
				row.day_of_week = calendar_row.day_of_week
				row.day_type = calendar_row.day_type
				row.day_team = ", ".join(day_teams)
				row.night_team = ", ".join(night_teams)
				row.off_teams = ", ".join(off_teams)

				if self.meta_has_child_field("Shift Plan Team Rotation", "cost_group"):
					row.cost_group = cost_group

		return self.as_dict()

	def expand_rotation_pattern_by_date_and_cost_group(self):
		rotation_by_pattern_day = self.get_rotation_pattern_map()
		rotation_map = {}

		anchor_date = getdate(getattr(self, "rotation_anchor_date", None) or self.calendar_start_date)
		pattern_days = cint(getattr(self, "rotation_pattern_days", 0)) or self.get_max_rotation_pattern_day()

		if not pattern_days:
			frappe.throw("Rotation Pattern Days must be set.")

		for calendar_row in self.shift_calendar:
			row_date = getdate(calendar_row.date)
			pattern_day = ((row_date - anchor_date).days % pattern_days) + 1

			rotation_map[row_date] = {}

			for cost_group, teams in rotation_by_pattern_day.items():
				rotation_map[row_date][cost_group] = {}

				for team_label in self.get_team_labels_for_cost_group(cost_group):
					assignment = teams.get(pattern_day, {}).get(team_label, "Off")
					rotation_map[row_date][cost_group][team_label] = assignment

			self.apply_sunday_rotation_rule(row_date, calendar_row, rotation_map)

		return rotation_map

	def apply_sunday_rotation_rule(self, row_date, calendar_row, rotation_map):
		rule = getattr(self, "sunday_rotation_rule", None) or "Follow Pattern"

		if rule != "Extend Saturday Day Team":
			return

		if calendar_row.day_type != "Sunday":
			return

		if not cint(calendar_row.day_shift_required):
			return

		previous_date = row_date - timedelta(days=1)

		if previous_date not in rotation_map:
			return

		for cost_group in rotation_map.get(row_date, {}):
			saturday_assignments = rotation_map.get(previous_date, {}).get(cost_group, {})
			sunday_assignments = rotation_map.get(row_date, {}).get(cost_group, {})

			saturday_day_teams = [
				team_label
				for team_label, assignment in saturday_assignments.items()
				if assignment == "Day"
			]

			if not saturday_day_teams:
				continue

			for team_label in sunday_assignments:
				if team_label in saturday_day_teams:
					sunday_assignments[team_label] = "Day"
				elif sunday_assignments[team_label] == "Day":
					sunday_assignments[team_label] = "Off"

	def get_rotation_pattern_map(self):
		result = defaultdict(lambda: defaultdict(dict))

		for row in self.rotation_pattern or []:
			cost_group = row.cost_group
			pattern_day = cint(row.pattern_day)
			team_label = row.team_label
			assignment = row.shift_assignment or "Off"

			if not cost_group or not pattern_day or not team_label:
				continue

			result[cost_group][pattern_day][team_label] = assignment

		return result

	def get_max_rotation_pattern_day(self):
		max_day = 0

		for row in self.rotation_pattern or []:
			max_day = max(max_day, cint(row.pattern_day))

		return max_day

	def get_team_labels_for_cost_group(self, cost_group):
		labels = []

		for row in self.rotation_pattern or []:
			if row.cost_group == cost_group and row.team_label and row.team_label not in labels:
				labels.append(row.team_label)

		if labels:
			return labels

		return [row.team_label for row in self.shift_team_labels or [] if row.team_label]

	def get_sa_public_holidays(self, start_date, end_date):
		years = list(range(start_date.year, end_date.year + 1))

		try:
			from holidays import country_holidays

			za_holidays = country_holidays("ZA", years=years)
			return {
				getdate(holiday_date): holiday_name
				for holiday_date, holiday_name in za_holidays.items()
				if start_date <= getdate(holiday_date) <= end_date
			}
		except Exception as exc:
			frappe.log_error(
				title="Shift Plan public holiday generation failed",
				message=frappe.get_traceback(),
			)

			frappe.throw(
				"Unable to generate South African public holidays. "
				"Please confirm that the Python 'holidays' package is installed in this site environment. "
				f"Original error: {exc}"
			)

	def get_day_type(self, current_date, public_holidays):
		if current_date in public_holidays:
			return "Public Holiday"

		if current_date.weekday() == 5:
			return "Saturday"

		if current_date.weekday() == 6:
			return "Sunday"

		return "Normal"

	def get_day_operating_values(self, day_name, day_type):
		if day_type == "Public Holiday":
			return self.get_public_holiday_operating_values(day_name)

		if day_type == "Saturday":
			return self.get_saturday_operating_values()

		if day_type == "Sunday":
			return self.get_sunday_operating_values()

		if not self.is_weekday_included(day_name):
			return 0, 0, 0

		return self.get_normal_day_operating_values()

	def is_weekday_included(self, day_name):
		weekday_map = {
			"Monday": self.include_mondays,
			"Tuesday": self.include_tuesdays,
			"Wednesday": self.include_wednesdays,
			"Thursday": self.include_thursdays,
			"Friday": self.include_fridays,
			"Saturday": self.include_saturdays,
			"Sunday": self.include_sundays,
		}

		return cint(weekday_map.get(day_name, 0)) == 1

	def get_default_shift_hours(self):
		return flt(self.default_shift_hours or 0)

	def get_default_day_shift_hours(self):
		day_shift_hours = flt(getattr(self, "default_day_shift_hours", 0))

		if day_shift_hours:
			return day_shift_hours

		return self.get_default_shift_hours()

	def get_default_operating_hours(self):
		operating_hours = flt(getattr(self, "default_operating_hours", 0))

		if operating_hours:
			return operating_hours

		if self.operating_model in ("24h Mon-Sat", "24h Mon-Sun"):
			return 24

		return self.get_default_shift_hours()

	def get_shift_requirements_from_operating_hours(self, operating_hours):
		operating_hours = flt(operating_hours)
		default_shift_hours = self.get_default_shift_hours()

		if operating_hours <= 0:
			return 0, 0, 0

		if default_shift_hours <= 0:
			if operating_hours >= 24:
				return operating_hours, 1, 1

			return operating_hours, 1, 0

		if operating_hours > default_shift_hours:
			return operating_hours, 1, 1

		return operating_hours, 1, 0

	def get_normal_day_operating_values(self):
		operating_hours = self.get_default_operating_hours()

		if self.operating_model == "Day Shift Only":
			operating_hours = self.get_default_day_shift_hours()

		return self.get_shift_requirements_from_operating_hours(operating_hours)

	def get_saturday_operating_values(self):
		if not cint(self.include_saturdays):
			return 0, 0, 0

		return self.get_special_day_work_model_values(self.saturday_work_model)

	def get_sunday_operating_values(self):
		if not cint(self.include_sundays):
			return 0, 0, 0

		return self.get_special_day_work_model_values(self.sunday_work_model)

	def get_public_holiday_operating_values(self, day_name):
		if not self.is_weekday_included(day_name):
			return 0, 0, 0

		operating_hours = self.get_default_operating_hours()

		if self.operating_model == "Day Shift Only":
			operating_hours = self.get_default_day_shift_hours()

		return self.get_shift_requirements_from_operating_hours(operating_hours)

	def get_special_day_work_model_values(self, work_model):
		if work_model == "No Work":
			return 0, 0, 0

		if work_model == "8h Day Only":
			return 8, 1, 0

		if work_model == "12h Day Only":
			return 12, 1, 0

		if work_model == "24h":
			return 24, 1, 1

		if work_model == "Custom":
			operating_hours = self.get_default_operating_hours()
			return self.get_shift_requirements_from_operating_hours(operating_hours)

		return 0, 0, 0

	@frappe.whitelist()
	def calculate_output(self):
		self.validate_dates()
		self.set_calendar_days()
		self.set_staffing_totals()

		if not self.shift_calendar:
			frappe.throw("Please generate the Shift Calendar before calculating output.")

		if not self.employees:
			frappe.throw("Please add Staffing rows before calculating output.")

		if self.requires_rotation_pattern() and not self.rotation_pattern:
			frappe.throw("Please set a Rotation Pattern before calculating output.")

		self.output = []

		if self.parent_has_field("team_output"):
			self.team_output = []

		detail_rows = []

		for staffing_row in self.employees:
			detail_row = self.calculate_staffing_output_row(staffing_row)
			detail_rows.append(detail_row)
			self.append("output", self.filter_child_row("Shift Plan Output", detail_row))

		for total_row in self.build_cost_group_totals(detail_rows):
			self.append("output", self.filter_child_row("Shift Plan Output", total_row))

		site_total = self.build_site_total(detail_rows)

		if site_total:
			self.append("output", self.filter_child_row("Shift Plan Output", site_total))

		if self.parent_has_field("team_output"):
			for team_row in self.build_exact_team_output():
				self.append("team_output", self.filter_child_row("Shift Plan Team Output", team_row))

		return self.as_dict()

	def requires_rotation_pattern(self):
		for row in self.employees or []:
			if row.working_hours in ("Shift Pattern", "Night Shift Only"):
				return True

		return False

	def calculate_staffing_output_row(self, staffing_row):
		employees_per_shift = flt(staffing_row.employees_per_shift)
		number_of_shift_teams = cint(staffing_row.number_of_shift_teams)
		total_employees = flt(staffing_row.total_employees)
		hourly_rate = flt(staffing_row.hourly_rate)
		fixed_costs = self.get_fixed_cost_for_staffing_row(staffing_row)

		required_shifts = 0
		normal_day_hours = 0
		sunday_hours = 0
		public_holiday_hours = 0

		for calendar_row in self.shift_calendar:
			if flt(calendar_row.planned_operating_hours) <= 0:
				continue

			if not self.is_calendar_row_allowed_for_staffing(calendar_row, staffing_row):
				continue

			shift_count = self.get_required_shift_count_for_staffing(calendar_row, staffing_row)

			if shift_count <= 0:
				continue

			hours_per_shift = self.get_hours_per_shift_for_staffing(calendar_row, staffing_row, shift_count)
			row_required_hours = employees_per_shift * shift_count * hours_per_shift

			required_shifts += employees_per_shift * shift_count

			hour_bucket = self.get_hour_bucket(calendar_row)

			if hour_bucket == "sunday":
				sunday_hours += row_required_hours
			elif hour_bucket == "public_holiday":
				public_holiday_hours += row_required_hours
			else:
				normal_day_hours += row_required_hours

		available_normal_hours = self.get_available_normal_hours(staffing_row)

		normal_hours = min(normal_day_hours, available_normal_hours)
		normal_overtime_hours = max(normal_day_hours - available_normal_hours, 0)

		if not cint(self.treat_sundays_as_overtime):
			normal_hours += sunday_hours
			sunday_hours = 0

		if not cint(self.treat_public_holidays_as_overtime):
			normal_hours += public_holiday_hours
			public_holiday_hours = 0

		total_overtime_hours = normal_overtime_hours + sunday_hours + public_holiday_hours
		required_hours = normal_day_hours + sunday_hours + public_holiday_hours

		normal_cost = normal_hours * hourly_rate
		normal_overtime_cost = normal_overtime_hours * hourly_rate * flt(self.normal_ot_multiplier)
		sunday_overtime_cost = sunday_hours * hourly_rate * flt(self.sunday_ot_multiplier)
		public_holiday_overtime_cost = public_holiday_hours * hourly_rate * flt(self.public_holiday_ot_multiplier)

		total_overtime_cost = normal_overtime_cost + sunday_overtime_cost + public_holiday_overtime_cost
		total_cost = normal_cost + total_overtime_cost + fixed_costs

		baseline_overtime_hours = normal_overtime_hours
		baseline_overtime_cost = normal_overtime_cost
		excluded_overtime_hours = 0
		excluded_overtime_cost = 0

		if cint(self.include_sunday_work_in_baseline):
			baseline_overtime_hours += sunday_hours
			baseline_overtime_cost += sunday_overtime_cost
		else:
			excluded_overtime_hours += sunday_hours
			excluded_overtime_cost += sunday_overtime_cost

		if cint(self.include_public_holiday_work_in_baseline):
			baseline_overtime_hours += public_holiday_hours
			baseline_overtime_cost += public_holiday_overtime_cost
		else:
			excluded_overtime_hours += public_holiday_hours
			excluded_overtime_cost += public_holiday_overtime_cost

		period_revenue = flt(self.period_revenue)

		overtime_percent_of_revenue = self.get_percent(total_overtime_cost, period_revenue)
		baseline_overtime_percent_of_revenue = self.get_percent(baseline_overtime_cost, period_revenue)
		overtime_percent_of_normal_time = self.get_percent(total_overtime_hours, normal_hours)

		comments = self.get_detail_comment(
			normal_day_hours=normal_day_hours,
			available_normal_hours=available_normal_hours,
			sunday_hours=sunday_hours,
			public_holiday_hours=public_holiday_hours,
			staffing_row=staffing_row,
		)

		return {
			"output_type": "Detail",
			"cost_group": staffing_row.cost_group,
			"designation": staffing_row.designation,
			"pay_basis": self.normalize_pay_basis(staffing_row.pay_basis),
			"working_hours": staffing_row.working_hours,
			"employees_per_shift": employees_per_shift,
			"number_of_shift_teams": number_of_shift_teams,
			"total_employees": total_employees,
			"required_shifts": required_shifts,
			"required_hours": required_hours,
			"normal_hours": normal_hours,
			"normal_overtime_hours": normal_overtime_hours,
			"sunday_hours": sunday_hours,
			"public_holiday_hours": public_holiday_hours,
			"total_overtime_hours": total_overtime_hours,
			"overtime_percent_of_normal_time": overtime_percent_of_normal_time,
			"hourly_rate": hourly_rate,
			"normal_cost": normal_cost,
			"normal_overtime_cost": normal_overtime_cost,
			"sunday_overtime_cost": sunday_overtime_cost,
			"public_holiday_overtime_cost": public_holiday_overtime_cost,
			"fixed_costs": fixed_costs,
			"total_cost": total_cost,
			"period_revenue": period_revenue,
			"overtime_percent_of_revenue": overtime_percent_of_revenue,
			"baseline_overtime_hours": baseline_overtime_hours,
			"baseline_overtime_cost": baseline_overtime_cost,
			"baseline_overtime_percent_of_revenue": baseline_overtime_percent_of_revenue,
			"excluded_overtime_hours": excluded_overtime_hours,
			"excluded_overtime_cost": excluded_overtime_cost,
			"comments": comments,
		}

	def get_fixed_cost_for_staffing_row(self, staffing_row):
		fixed_costs = flt(staffing_row.fixed_costs)

		if not self.meta_has_child_field("Shift Plan Staffing", "fixed_cost_basis"):
			return fixed_costs

		if getattr(staffing_row, "fixed_cost_basis", None) == "Monthly":
			return fixed_costs * self.get_period_factor_for_pay_basis(staffing_row.pay_basis)

		return fixed_costs

	def get_available_normal_hours(self, staffing_row):
		total_employees = flt(staffing_row.total_employees)
		normal_hours_limit = flt(self.normal_hours_limit)

		if self.normal_hours_limit_basis == "Per Calendar Period":
			return total_employees * normal_hours_limit

		period_factor = self.get_period_factor_for_pay_basis(staffing_row.pay_basis)
		return total_employees * normal_hours_limit * period_factor

	def get_period_factor_for_pay_basis(self, pay_basis):
		buckets = self.get_period_buckets_for_pay_basis(pay_basis)
		factor = 0

		for bucket in buckets:
			selected_days = (bucket["end_date"] - bucket["start_date"]).days + 1
			full_days = (bucket["full_end_date"] - bucket["full_start_date"]).days + 1

			if full_days <= 0:
				continue

			factor += selected_days / full_days

		return factor or 1

	def is_calendar_row_allowed_for_staffing(self, calendar_row, staffing_row):
		working_days = getattr(staffing_row, "working_days", None) or self.get_default_working_days_for_staffing(staffing_row)
		return self.is_day_name_allowed_by_working_days(calendar_row.day_of_week, working_days)

	def get_default_working_days_for_staffing(self, staffing_row):
		if getattr(staffing_row, "working_hours", None) == "Day Shift Only":
			return "Monday to Friday"

		return "Follow Shift Plan Calendar"

	def get_required_shift_count_for_staffing(self, calendar_row, staffing_row):
		working_hours = staffing_row.working_hours

		day_required = cint(calendar_row.day_shift_required)
		night_required = cint(calendar_row.night_shift_required)

		if working_hours == "Shift Pattern":
			return day_required + night_required

		if working_hours == "Day Shift Only":
			return day_required

		if working_hours == "Night Shift Only":
			return night_required

		return 0

	def get_hours_per_shift_for_staffing(self, calendar_row, staffing_row, shift_count):
		if staffing_row.working_hours == "Day Shift Only":
			planned_operating_hours = flt(calendar_row.planned_operating_hours)
			day_shift_hours = self.get_default_day_shift_hours()

			if planned_operating_hours <= 0:
				return 0

			return min(planned_operating_hours, day_shift_hours)

		return self.get_hours_per_shift(calendar_row, shift_count)

	def get_hours_per_shift(self, calendar_row, shift_count):
		planned_operating_hours = flt(calendar_row.planned_operating_hours)
		default_shift_hours = self.get_default_shift_hours()

		if shift_count <= 0:
			return 0

		if planned_operating_hours <= 0:
			return 0

		if shift_count == 1:
			if default_shift_hours:
				return min(planned_operating_hours, default_shift_hours)

			return planned_operating_hours

		return planned_operating_hours / shift_count

	def get_hour_bucket(self, calendar_row):
		if calendar_row.day_type == "Public Holiday" and cint(self.treat_public_holidays_as_overtime):
			return "public_holiday"

		if calendar_row.day_type == "Sunday" and cint(self.treat_sundays_as_overtime):
			return "sunday"

		return "normal"

	def build_exact_team_output(self):
		rows = []

		patterns = self.get_unique_team_output_patterns()
		rotation_map = self.expand_rotation_pattern_by_date_and_cost_group()

		for pattern in patterns:
			buckets = self.get_period_buckets_for_pay_basis(pattern["pay_basis"])

			for period in buckets:
				rows.extend(self.build_team_output_for_pattern_and_period(pattern, period, rotation_map))

		rows.extend(self.build_average_team_output_rows(rows))

		return rows

	def get_unique_team_output_patterns(self):
		patterns = {}

		for staffing_row in self.employees:
			cost_group = staffing_row.cost_group
			pay_basis = self.normalize_pay_basis(staffing_row.pay_basis)
			working_hours = staffing_row.working_hours
			working_days = getattr(staffing_row, "working_days", None) or self.get_default_working_days_for_staffing(staffing_row)

			key = (cost_group, pay_basis, working_hours, working_days)

			if key not in patterns:
				patterns[key] = {
					"cost_group": cost_group,
					"pay_basis": pay_basis,
					"working_hours": working_hours,
					"working_days": working_days,
				}

		return list(patterns.values())

	def build_team_output_for_pattern_and_period(self, pattern, period, rotation_map):
		raw = defaultdict(lambda: {"normal": 0, "sunday": 0, "public_holiday": 0})

		for calendar_row in self.shift_calendar:
			row_date = getdate(calendar_row.date)

			if row_date < period["start_date"] or row_date > period["end_date"]:
				continue

			if flt(calendar_row.planned_operating_hours) <= 0:
				continue

			if not self.is_day_name_allowed_by_working_days(calendar_row.day_of_week, pattern["working_days"]):
				continue

			hour_bucket = self.get_hour_bucket(calendar_row)
			working_hours = pattern["working_hours"]
			cost_group = pattern["cost_group"]

			if working_hours == "Shift Pattern":
				assignments = rotation_map.get(row_date, {}).get(cost_group, {})

				for team_label, assignment in assignments.items():
					if assignment == "Day" and cint(calendar_row.day_shift_required):
						raw[team_label][hour_bucket] += self.get_day_shift_hours_for_calendar_row(calendar_row)

					if assignment == "Night" and cint(calendar_row.night_shift_required):
						raw[team_label][hour_bucket] += self.get_night_shift_hours_for_calendar_row(calendar_row)

			elif working_hours == "Day Shift Only":
				if cint(calendar_row.day_shift_required):
					raw["Day Shift Only"][hour_bucket] += self.get_default_day_shift_hours()

			elif working_hours == "Night Shift Only":
				assignments = rotation_map.get(row_date, {}).get(cost_group, {})

				for team_label, assignment in assignments.items():
					if assignment == "Night" and cint(calendar_row.night_shift_required):
						raw[team_label][hour_bucket] += self.get_night_shift_hours_for_calendar_row(calendar_row)

		output_rows = []

		for team_label, values in raw.items():
			raw_normal_hours = flt(values["normal"])
			sunday_hours = flt(values["sunday"])
			public_holiday_hours = flt(values["public_holiday"])

			period_normal_limit = self.get_normal_hours_limit_for_period(period)

			ordinary_hours = min(raw_normal_hours, period_normal_limit)
			normal_overtime_hours = max(raw_normal_hours - period_normal_limit, 0)
			total_overtime_hours = normal_overtime_hours + sunday_hours + public_holiday_hours
			total_hours = ordinary_hours + total_overtime_hours

			output_rows.append(
				{
					"period_label": period["label"],
					"period_start_date": period["start_date"],
					"period_end_date": period["end_date"],
					"cost_group": pattern["cost_group"],
					"pay_basis": pattern["pay_basis"],
					"team_label": team_label,
					"working_hours": pattern["working_hours"],
					"working_days": pattern["working_days"],
					"ordinary_hours": ordinary_hours,
					"normal_overtime_hours": normal_overtime_hours,
					"sunday_hours": sunday_hours,
					"public_holiday_hours": public_holiday_hours,
					"total_overtime_hours": total_overtime_hours,
					"total_hours": total_hours,
					"overtime_percent_of_normal_time": self.get_percent(total_overtime_hours, ordinary_hours),
					"comments": "Calculated from Rotation Pattern, calendar and pay-period buckets.",
				}
			)

		return output_rows

	def build_average_team_output_rows(self, rows):
		period_rows = [row for row in rows if row.get("period_label") != "Average"]

		if not period_rows:
			return []

		grouped = defaultdict(list)

		for row in period_rows:
			key = (
				row.get("cost_group"),
				row.get("pay_basis"),
				row.get("team_label"),
				row.get("working_hours"),
				row.get("working_days"),
			)
			grouped[key].append(row)

		average_rows = []

		for key, grouped_rows in grouped.items():
			cost_group, pay_basis, team_label, working_hours, working_days = key
			count = len(grouped_rows)

			if not count:
				continue

			ordinary_hours = sum(flt(row.get("ordinary_hours")) for row in grouped_rows) / count
			normal_overtime_hours = sum(flt(row.get("normal_overtime_hours")) for row in grouped_rows) / count
			sunday_hours = sum(flt(row.get("sunday_hours")) for row in grouped_rows) / count
			public_holiday_hours = sum(flt(row.get("public_holiday_hours")) for row in grouped_rows) / count
			total_overtime_hours = normal_overtime_hours + sunday_hours + public_holiday_hours
			total_hours = ordinary_hours + total_overtime_hours

			average_rows.append(
				{
					"period_label": "Average",
					"period_start_date": None,
					"period_end_date": None,
					"cost_group": cost_group,
					"pay_basis": pay_basis,
					"team_label": team_label,
					"working_hours": working_hours,
					"working_days": working_days,
					"ordinary_hours": ordinary_hours,
					"normal_overtime_hours": normal_overtime_hours,
					"sunday_hours": sunday_hours,
					"public_holiday_hours": public_holiday_hours,
					"total_overtime_hours": total_overtime_hours,
					"total_hours": total_hours,
					"overtime_percent_of_normal_time": self.get_percent(total_overtime_hours, ordinary_hours),
					"comments": f"Average across {count} pay period(s).",
				}
			)

		return average_rows

	def get_normal_hours_limit_for_period(self, period):
		if self.normal_hours_limit_basis == "Per Calendar Period":
			return flt(self.normal_hours_limit)

		selected_days = (period["end_date"] - period["start_date"]).days + 1
		full_days = (period["full_end_date"] - period["full_start_date"]).days + 1

		if full_days <= 0:
			return flt(self.normal_hours_limit)

		return flt(self.normal_hours_limit) * selected_days / full_days

	def get_period_buckets_for_pay_basis(self, pay_basis):
		pay_basis = self.normalize_pay_basis(pay_basis)

		if pay_basis == "Hourly":
			return self.get_hourly_pay_period_buckets()

		if pay_basis == "Salaried":
			return self.get_salaried_pay_period_buckets()

		return self.get_hourly_pay_period_buckets()

	def get_hourly_pay_period_buckets(self):
		start_date = getdate(self.calendar_start_date)
		end_date = getdate(self.calendar_end_date)

		start_day = cint(self.hourly_pay_period_start_day or 16)
		end_day = cint(self.hourly_pay_period_end_day or 15)

		full_start = self.get_hourly_period_start_for_date(start_date, start_day)

		buckets = []

		while full_start <= end_date:
			full_end = self.get_hourly_period_end(full_start, end_day)

			selected_start = max(start_date, full_start)
			selected_end = min(end_date, full_end)

			if selected_start <= selected_end:
				buckets.append(
					{
						"label": full_end.strftime("%B %Y"),
						"start_date": selected_start,
						"end_date": selected_end,
						"full_start_date": full_start,
						"full_end_date": full_end,
					}
				)

			full_start = self.add_month_preserve_day(full_start, start_day)

		return buckets

	def get_hourly_period_start_for_date(self, input_date, start_day):
		input_date = getdate(input_date)

		if input_date.day >= start_day:
			return self.safe_date(input_date.year, input_date.month, start_day)

		previous_month = self.add_month_preserve_day(input_date.replace(day=1), 1, months=-1)
		return self.safe_date(previous_month.year, previous_month.month, start_day)

	def get_hourly_period_end(self, full_start, end_day):
		next_month_start = self.add_month_preserve_day(full_start, 1)
		return self.safe_date(next_month_start.year, next_month_start.month, end_day)

	def get_salaried_pay_period_buckets(self):
		start_date = getdate(self.calendar_start_date)
		end_date = getdate(self.calendar_end_date)

		full_start = start_date.replace(day=1)

		buckets = []

		while full_start <= end_date:
			full_end = self.add_month(full_start) - timedelta(days=1)

			selected_start = max(start_date, full_start)
			selected_end = min(end_date, full_end)

			if selected_start <= selected_end:
				buckets.append(
					{
						"label": full_start.strftime("%B %Y"),
						"start_date": selected_start,
						"end_date": selected_end,
						"full_start_date": full_start,
						"full_end_date": full_end,
					}
				)

			full_start = self.add_month(full_start)

		return buckets

	def add_month(self, input_date):
		if input_date.month == 12:
			return input_date.replace(year=input_date.year + 1, month=1, day=1)

		return input_date.replace(month=input_date.month + 1, day=1)

	def add_month_preserve_day(self, input_date, day, months=1):
		year = input_date.year
		month = input_date.month + months

		while month > 12:
			year += 1
			month -= 12

		while month < 1:
			year -= 1
			month += 12

		return self.safe_date(year, month, day)

	def safe_date(self, year, month, day):
		last_day = monthrange(year, month)[1]
		return getdate(f"{year}-{month:02d}-{min(cint(day), last_day):02d}")

	def is_day_name_allowed_by_working_days(self, day_name, working_days):
		if working_days == "Follow Shift Plan Calendar":
			return True

		if working_days == "Monday to Friday":
			return day_name in ("Monday", "Tuesday", "Wednesday", "Thursday", "Friday")

		if working_days == "Monday to Saturday":
			return day_name in ("Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday")

		if working_days == "Monday to Sunday":
			return day_name in ("Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday")

		return True

	def get_day_shift_hours_for_calendar_row(self, calendar_row):
		if not cint(calendar_row.day_shift_required):
			return 0

		if cint(calendar_row.night_shift_required):
			return self.get_default_shift_hours()

		return min(flt(calendar_row.planned_operating_hours), self.get_default_shift_hours())

	def get_night_shift_hours_for_calendar_row(self, calendar_row):
		if not cint(calendar_row.night_shift_required):
			return 0

		if cint(calendar_row.day_shift_required):
			return self.get_default_shift_hours()

		return min(flt(calendar_row.planned_operating_hours), self.get_default_shift_hours())

	@frappe.whitelist()
	def get_visual_calendar_data(self):
		if not self.shift_calendar:
			return {"months": []}

		if not self.rotation_pattern:
			return {"months": []}

		rotation_map = self.expand_rotation_pattern_by_date_and_cost_group()
		months = []
		month_map = {}

		for calendar_row in self.shift_calendar:
			row_date = getdate(calendar_row.date)
			month_key = row_date.strftime("%Y-%m")
			month_label = row_date.strftime("%B %Y")

			if month_key not in month_map:
				month = {
					"key": month_key,
					"label": month_label,
					"days": [],
				}
				month_map[month_key] = month
				months.append(month)

			day_data = {
				"date": row_date.isoformat(),
				"day": row_date.day,
				"weekday": row_date.strftime("%A"),
				"day_type": calendar_row.day_type,
				"planned_operating_hours": flt(calendar_row.planned_operating_hours),
				"day_shift_required": cint(calendar_row.day_shift_required),
				"night_shift_required": cint(calendar_row.night_shift_required),
				"cost_groups": {},
			}

			for cost_group, assignments in rotation_map.get(row_date, {}).items():
				group_data = {
					"day": [],
					"night": [],
					"off": [],
					"day_hours": self.get_day_shift_hours_for_calendar_row(calendar_row) if cint(calendar_row.day_shift_required) else 0,
					"night_hours": self.get_night_shift_hours_for_calendar_row(calendar_row) if cint(calendar_row.night_shift_required) else 0,
				}

				for team_label, assignment in assignments.items():
					if assignment == "Day" and cint(calendar_row.day_shift_required):
						group_data["day"].append(team_label)
					elif assignment == "Night" and cint(calendar_row.night_shift_required):
						group_data["night"].append(team_label)
					else:
						group_data["off"].append(team_label)

				group_data["day_hours_label"] = f"({group_data['day_hours']:.0f}h)" if group_data["day"] else ""
				group_data["night_hours_label"] = f"({group_data['night_hours']:.0f}h)" if group_data["night"] else ""

				day_data["cost_groups"][cost_group] = group_data

			self.add_day_shift_only_cost_groups_to_visual_day(day_data, calendar_row)
			month_map[month_key]["days"].append(day_data)

		for month in months:
			month["weeks"] = self.build_month_weeks(month["days"])

		return {"months": months}

	def add_day_shift_only_cost_groups_to_visual_day(self, day_data, calendar_row):
		for staffing_row in self.employees or []:
			if staffing_row.working_hours != "Day Shift Only":
				continue

			cost_group = staffing_row.cost_group

			if not cost_group:
				continue

			if cost_group in day_data["cost_groups"]:
				continue

			allowed = self.is_calendar_row_allowed_for_staffing(calendar_row, staffing_row)
			day_required = cint(calendar_row.day_shift_required)

			group_data = {
				"day": [],
				"night": [],
				"off": [],
				"day_hours": self.get_default_day_shift_hours() if allowed and day_required else 0,
				"night_hours": 0,
			}

			if allowed and day_required:
				group_data["day"] = ["Day Shift Only"]
			else:
				group_data["off"] = ["Day Shift Only"]

			group_data["day_hours_label"] = f"({group_data['day_hours']:.0f}h)" if group_data["day"] else ""
			group_data["night_hours_label"] = ""

			day_data["cost_groups"][cost_group] = group_data

	def build_month_weeks(self, days):
		if not days:
			return []

		weeks = []
		current_week = [None, None, None, None, None, None, None]

		first_date = getdate(days[0]["date"])
		first_weekday = first_date.weekday()

		for index in range(first_weekday):
			current_week[index] = None

		for day in days:
			row_date = getdate(day["date"])
			weekday = row_date.weekday()

			if weekday == 0 and any(current_week):
				weeks.append(current_week)
				current_week = [None, None, None, None, None, None, None]

			current_week[weekday] = day

		if any(current_week):
			weeks.append(current_week)

		return weeks

	def get_detail_comment(self, normal_day_hours, available_normal_hours, sunday_hours, public_holiday_hours, staffing_row):
		comments = []

		if normal_day_hours > available_normal_hours:
			comments.append("Normal roster hours exceed available normal hours.")

		if sunday_hours and not cint(self.include_sunday_work_in_baseline):
			comments.append("Sunday overtime excluded from baseline.")

		if public_holiday_hours and not cint(self.include_public_holiday_work_in_baseline):
			comments.append("Public holiday overtime excluded from baseline.")

		if getattr(staffing_row, "working_days", None) and staffing_row.working_days != "Follow Shift Plan Calendar":
			comments.append(f"Working days limited to {staffing_row.working_days}.")

		if getattr(staffing_row, "fixed_cost_basis", None):
			comments.append(f"Fixed cost basis: {staffing_row.fixed_cost_basis}.")

		if not comments:
			comments.append("Calculated from approved Shift Plan calendar and staffing assumptions.")

		return " ".join(comments)

	def build_cost_group_totals(self, detail_rows):
		grouped = defaultdict(list)

		for row in detail_rows:
			grouped[row.get("cost_group") or "Unspecified"].append(row)

		total_rows = []

		for cost_group, rows in grouped.items():
			total_row = self.sum_output_rows(rows)
			total_row.update(
				{
					"output_type": "Cost Group Total",
					"cost_group": cost_group,
					"designation": None,
					"pay_basis": None,
					"working_hours": None,
					"comments": f"Total for {cost_group}.",
				}
			)

			total_rows.append(total_row)

		return total_rows

	def build_site_total(self, detail_rows):
		if not detail_rows:
			return None

		total_row = self.sum_output_rows(detail_rows)
		total_row.update(
			{
				"output_type": "Site Total",
				"cost_group": "Total",
				"designation": None,
				"pay_basis": None,
				"working_hours": None,
				"comments": "Total for Shift Plan.",
			}
		)

		return total_row

	def sum_output_rows(self, rows):
		period_revenue = flt(self.period_revenue)

		total = {
			"employees_per_shift": sum(flt(row.get("employees_per_shift")) for row in rows),
			"number_of_shift_teams": 0,
			"total_employees": sum(flt(row.get("total_employees")) for row in rows),
			"required_shifts": sum(flt(row.get("required_shifts")) for row in rows),
			"required_hours": sum(flt(row.get("required_hours")) for row in rows),
			"normal_hours": sum(flt(row.get("normal_hours")) for row in rows),
			"normal_overtime_hours": sum(flt(row.get("normal_overtime_hours")) for row in rows),
			"sunday_hours": sum(flt(row.get("sunday_hours")) for row in rows),
			"public_holiday_hours": sum(flt(row.get("public_holiday_hours")) for row in rows),
			"total_overtime_hours": sum(flt(row.get("total_overtime_hours")) for row in rows),
			"hourly_rate": 0,
			"normal_cost": sum(flt(row.get("normal_cost")) for row in rows),
			"normal_overtime_cost": sum(flt(row.get("normal_overtime_cost")) for row in rows),
			"sunday_overtime_cost": sum(flt(row.get("sunday_overtime_cost")) for row in rows),
			"public_holiday_overtime_cost": sum(flt(row.get("public_holiday_overtime_cost")) for row in rows),
			"fixed_costs": sum(flt(row.get("fixed_costs")) for row in rows),
			"total_cost": sum(flt(row.get("total_cost")) for row in rows),
			"period_revenue": period_revenue,
			"baseline_overtime_hours": sum(flt(row.get("baseline_overtime_hours")) for row in rows),
			"baseline_overtime_cost": sum(flt(row.get("baseline_overtime_cost")) for row in rows),
			"excluded_overtime_hours": sum(flt(row.get("excluded_overtime_hours")) for row in rows),
			"excluded_overtime_cost": sum(flt(row.get("excluded_overtime_cost")) for row in rows),
		}

		total_overtime_cost = (
			flt(total["normal_overtime_cost"])
			+ flt(total["sunday_overtime_cost"])
			+ flt(total["public_holiday_overtime_cost"])
		)

		total["overtime_percent_of_revenue"] = self.get_percent(total_overtime_cost, period_revenue)
		total["baseline_overtime_percent_of_revenue"] = self.get_percent(total["baseline_overtime_cost"], period_revenue)
		total["overtime_percent_of_normal_time"] = self.get_percent(
			total["total_overtime_hours"],
			total["normal_hours"],
		)

		return total

	def normalize_pay_basis(self, pay_basis):
		if pay_basis in ("Salary", "Salaried"):
			return "Salaried"

		if pay_basis == "Hourly":
			return "Hourly"

		if self.pay_period_type in ("Salary", "Salaried"):
			return "Salaried"

		if self.pay_period_type == "Hourly":
			return "Hourly"

		return "Hourly"

	def parent_has_field(self, fieldname):
		return self.meta.has_field(fieldname)

	def meta_has_child_field(self, doctype, fieldname):
		try:
			return frappe.get_meta(doctype).has_field(fieldname)
		except Exception:
			return False

	def filter_child_row(self, doctype, row):
		meta = frappe.get_meta(doctype)
		allowed_fields = {df.fieldname for df in meta.fields}
		return {key: value for key, value in row.items() if key in allowed_fields}

	def get_percent(self, value, base):
		value = flt(value)
		base = flt(base)

		if not base:
			return 0

		return (value / base) * 100

	@frappe.whitelist()
	def export_shift_plan_xlsx(self):
		"""Export a single-sheet workbook in the same practical layout as the original roster analysis.

		Important:
		- The calculation engine is not changed here. This method calls calculate_output() first.
		- The workbook has one sheet only: Sheet1.
		- Date cells are real Excel date cells using Excel's built-in date format 14.
		- The Day row is an Excel formula based on the Date row.
		"""
		self.calculate_output()

		try:
			from openpyxl import Workbook
		except Exception as exc:
			frappe.throw(f"Unable to export XLSX because openpyxl is not available: {exc}")

		wb = Workbook()
		ws = wb.active
		ws.title = "Sheet1"

		self.add_original_style_shift_roster_sheet(ws)

		output = BytesIO()
		wb.save(output)
		output.seek(0)

		filename = f"{self.name} - Shift Roster Analysis.xlsx"

		from frappe.utils.file_manager import save_file

		file_doc = save_file(
			fname=filename,
			content=output.getvalue(),
			dt=self.doctype,
			dn=self.name,
			is_private=1,
		)

		return {
			"file_url": file_doc.file_url,
			"file_name": filename,
		}

	def add_original_style_shift_roster_sheet(self, ws):
		from openpyxl.styles import Font
		from openpyxl.styles import numbers
		from openpyxl.utils import get_column_letter

		date_format = numbers.FORMAT_DATE_XLSX14
		number_format = "0.00"
		currency_format = '"R"#,##0.00'
		percent_format = "0.00%"

		calendar_rows = sorted(list(self.shift_calendar or []), key=lambda row: getdate(row.date))
		visual_data = self.get_visual_calendar_data()
		visual_by_date = {}

		for month in visual_data.get("months", []):
			for day in month.get("days", []):
				if day.get("date"):
					visual_by_date[getdate(day.get("date"))] = day

		date_row = 1
		pph_row = 2
		day_row = 3
		first_date_col = 2

		ws.cell(row=date_row, column=1, value="Date")
		ws.cell(row=pph_row, column=1, value="PPH")
		ws.cell(row=day_row, column=1, value="Day")

		date_to_col = {}

		for offset, calendar_row in enumerate(calendar_rows):
			col = first_date_col + offset
			row_date = getdate(calendar_row.date)
			date_to_col[row_date] = col

			date_cell = ws.cell(row=date_row, column=col, value=row_date)
			date_cell.number_format = date_format

			ws.cell(
				row=pph_row,
				column=col,
				value="PPH" if calendar_row.day_type == "Public Holiday" else None,
			)

			date_ref = f"{get_column_letter(col)}${date_row}"
			day_cell = ws.cell(row=day_row, column=col, value=f'=TEXT({date_ref},"dddd")')
			day_cell.number_format = "General"

		team_rows_start = 4
		team_row_map = {}
		current_row = team_rows_start

		for cost_group, team_label in self.get_export_cost_group_team_pairs(visual_by_date):
			ws.cell(row=current_row, column=1, value=f"{cost_group} - {team_label}")
			team_row_map[(cost_group, team_label)] = current_row

			for row_date, col in date_to_col.items():
				day_data = visual_by_date.get(row_date, {})
				group_data = day_data.get("cost_groups", {}).get(cost_group, {})
				value = None

				if team_label in (group_data.get("day") or []):
					value = flt(group_data.get("day_hours") or 0)
				elif team_label in (group_data.get("night") or []):
					value = flt(group_data.get("night_hours") or 0)

				cell = ws.cell(row=current_row, column=col, value=value)
				cell.number_format = number_format

			current_row += 1

		summary_start_row = current_row + 2
		self.write_original_style_team_summary(
			ws=ws,
			start_row=summary_start_row,
			date_to_col=date_to_col,
			team_row_map=team_row_map,
			date_row=date_row,
			pph_row=pph_row,
			day_row=day_row,
			number_format=number_format,
			percent_format=percent_format,
		)

		labour_start_row = self.get_last_used_row(ws) + 3
		self.write_original_style_labour_summary(
			ws=ws,
			start_row=labour_start_row,
			currency_format=currency_format,
			percent_format=percent_format,
		)

		for col in range(1, ws.max_column + 1):
			ws.column_dimensions[get_column_letter(col)].width = 13 if col > 1 else 24

		for row in range(1, ws.max_row + 1):
			ws.row_dimensions[row].height = 18

		for cell in ws[date_row]:
			if cell.column >= first_date_col:
				cell.number_format = date_format

		# Minimal emphasis only; no fills, borders, merged cells or dashboard styling.
		for row in (date_row, pph_row, day_row):
			ws.cell(row=row, column=1).font = Font(bold=True)

	def get_export_cost_group_team_pairs(self, visual_by_date):
		pairs = []
		seen = set()

		for row in self.team_output or []:
			cost_group = row.cost_group or "Unspecified"
			team_label = row.team_label or "Unspecified"
			key = (cost_group, team_label)

			if key not in seen:
				seen.add(key)
				pairs.append(key)

		for day_data in visual_by_date.values():
			for cost_group, group_data in (day_data.get("cost_groups") or {}).items():
				for team_label in (group_data.get("day") or []) + (group_data.get("night") or []) + (group_data.get("off") or []):
					key = (cost_group or "Unspecified", team_label or "Unspecified")

					if key not in seen:
						seen.add(key)
						pairs.append(key)

		return sorted(pairs, key=lambda pair: (pair[0] or "", self.get_team_sort_value(pair[1])))

	def get_team_sort_value(self, team_label):
		team_label = str(team_label or "")

		if team_label.lower().startswith("team "):
			return team_label[5:].strip().upper()

		return team_label.upper()

	def write_original_style_team_summary(
		self,
		ws,
		start_row,
		date_to_col,
		team_row_map,
		date_row,
		pph_row,
		day_row,
		number_format,
		percent_format,
	):
		from openpyxl.styles import Font
		from openpyxl.utils import get_column_letter

		headers = [
			"Period",
			"Cost Group",
			"Team",
			"NT",
			"OT 1.5",
			"Sunday",
			"PPH",
			"Total OT",
			"Total Hours",
			"OT % NT",
		]

		for col_idx, header in enumerate(headers, start=1):
			ws.cell(row=start_row, column=col_idx, value=header)
			ws.cell(row=start_row, column=col_idx).font = Font(bold=True)

		rows = sorted(
			list(self.team_output or []),
			key=lambda row: (
				1 if row.period_label == "Average" else 0,
				getdate(row.period_start_date or "2999-12-31").toordinal() if row.period_label != "Average" else 99999999,
				row.cost_group or "",
				self.get_team_sort_value(row.team_label),
			),
		)

		summary_rows_by_key = defaultdict(list)
		current_row = start_row + 1

		for output_row in rows:
			cost_group = output_row.cost_group or "Unspecified"
			team_label = output_row.team_label or "Unspecified"
			team_roster_row = team_row_map.get((cost_group, team_label))

			ws.cell(row=current_row, column=1, value=output_row.period_label)
			ws.cell(row=current_row, column=2, value=cost_group)
			ws.cell(row=current_row, column=3, value=team_label)

			if output_row.period_label == "Average":
				average_source_rows = summary_rows_by_key.get((cost_group, team_label), [])

				if average_source_rows:
					for col_idx in range(4, 10):
						col_letter = get_column_letter(col_idx)
						refs = ",".join(f"{col_letter}{row_number}" for row_number in average_source_rows)
						ws.cell(row=current_row, column=col_idx, value=f"=AVERAGE({refs})")
				else:
					self.write_team_summary_values(ws, current_row, output_row)
			else:
				period_start = getdate(output_row.period_start_date)
				period_end = getdate(output_row.period_end_date)
				period_cols = [
					col for row_date, col in sorted(date_to_col.items())
					if period_start <= row_date <= period_end
				]

				if team_roster_row and period_cols:
					first_col = get_column_letter(min(period_cols))
					last_col = get_column_letter(max(period_cols))
					team_range = f"{first_col}{team_roster_row}:{last_col}{team_roster_row}"
					day_range = f"{first_col}${day_row}:{last_col}${day_row}"
					pph_range = f"{first_col}${pph_row}:{last_col}${pph_row}"
					normal_limit = flt(self.normal_hours_limit or 195)
					normal_formula = f'SUMIFS({team_range},{day_range},"<>Sunday",{pph_range},"<>PPH")'

					ws.cell(row=current_row, column=4, value=f"=MIN({normal_formula},{normal_limit})")
					ws.cell(row=current_row, column=5, value=f"=MAX({normal_formula}-{normal_limit},0)")
					ws.cell(row=current_row, column=6, value=f'=SUMIFS({team_range},{day_range},"Sunday")')
					ws.cell(row=current_row, column=7, value=f'=SUMIFS({team_range},{pph_range},"PPH")')
					ws.cell(row=current_row, column=8, value=f"=E{current_row}+F{current_row}+G{current_row}")
					ws.cell(row=current_row, column=9, value=f"=D{current_row}+H{current_row}")
					ws.cell(row=current_row, column=10, value=f'=IF(D{current_row}=0,"",H{current_row}/D{current_row})')
					summary_rows_by_key[(cost_group, team_label)].append(current_row)
				else:
					self.write_team_summary_values(ws, current_row, output_row)

			for col_idx in range(4, 10):
				ws.cell(row=current_row, column=col_idx).number_format = percent_format if col_idx == 10 else number_format

			current_row += 1

	def write_team_summary_values(self, ws, row_number, output_row):
		ws.cell(row=row_number, column=4, value=flt(output_row.ordinary_hours))
		ws.cell(row=row_number, column=5, value=flt(output_row.normal_overtime_hours))
		ws.cell(row=row_number, column=6, value=flt(output_row.sunday_hours))
		ws.cell(row=row_number, column=7, value=flt(output_row.public_holiday_hours))
		ws.cell(row=row_number, column=8, value=flt(output_row.total_overtime_hours))
		ws.cell(row=row_number, column=9, value=flt(output_row.total_hours))
		ws.cell(
			row=row_number,
			column=10,
			value=flt(output_row.overtime_percent_of_normal_time) / 100 if flt(output_row.overtime_percent_of_normal_time) else 0,
		)

	def write_original_style_labour_summary(self, ws, start_row, currency_format, percent_format):
		from openpyxl.styles import Font

		site_total = None

		for output_row in self.output or []:
			if output_row.output_type == "Site Total":
				site_total = output_row
				break

		if not site_total:
			return

		revenue = flt(site_total.period_revenue or self.period_revenue)
		normal_time_cost = flt(site_total.normal_cost)
		fixed_labour_cost = flt(site_total.fixed_costs)
		normal_overtime_cost = flt(site_total.normal_overtime_cost)
		sunday_cost = flt(site_total.sunday_overtime_cost)
		public_holiday_cost = flt(site_total.public_holiday_overtime_cost)

		ws.cell(row=start_row, column=1, value="Labour Cost Summary")
		ws.cell(row=start_row, column=1).font = Font(bold=True)

		header_row = start_row + 1
		ws.cell(row=header_row, column=1, value="Description")
		ws.cell(row=header_row, column=2, value="Amount")
		ws.cell(row=header_row, column=3, value="% of Revenue")

		for col_idx in range(1, 4):
			ws.cell(row=header_row, column=col_idx).font = Font(bold=True)

		rows = [
			("Revenue / Turnover", revenue),
			("Normal Time Cost", normal_time_cost),
			("Fixed Labour Cost", fixed_labour_cost),
			("Overtime Cost", normal_overtime_cost),
			("Sunday Cost", sunday_cost),
			("Public Holiday Cost", public_holiday_cost),
			("Total Overtime Cost", None),
			("Total Labour Cost", None),
		]

		current_row = header_row + 1
		revenue_row = current_row
		normal_ot_row = current_row + 3
		sunday_row = current_row + 4
		pph_row = current_row + 5
		total_ot_row = current_row + 6
		normal_time_row = current_row + 1
		fixed_row = current_row + 2

		for label, amount in rows:
			ws.cell(row=current_row, column=1, value=label)

			if label == "Total Overtime Cost":
				ws.cell(row=current_row, column=2, value=f"=B{normal_ot_row}+B{sunday_row}+B{pph_row}")
			elif label == "Total Labour Cost":
				ws.cell(row=current_row, column=2, value=f"=B{normal_time_row}+B{fixed_row}+B{total_ot_row}")
			else:
				ws.cell(row=current_row, column=2, value=amount)

			if label == "Revenue / Turnover":
				ws.cell(row=current_row, column=3, value=None)
			else:
				ws.cell(row=current_row, column=3, value=f'=IF($B${revenue_row}=0,"",B{current_row}/$B${revenue_row})')

			ws.cell(row=current_row, column=2).number_format = currency_format
			ws.cell(row=current_row, column=3).number_format = percent_format

			if label in ("Total Overtime Cost", "Total Labour Cost"):
				for col_idx in range(1, 4):
					ws.cell(row=current_row, column=col_idx).font = Font(bold=True)

			current_row += 1

	def get_last_used_row(self, ws):
		last_row = 1

		for row in ws.iter_rows():
			for cell in row:
				if cell.value not in (None, ""):
					last_row = max(last_row, cell.row)

		return last_row
