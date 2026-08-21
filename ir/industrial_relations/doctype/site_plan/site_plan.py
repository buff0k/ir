# Copyright (c) 2026, BuFf0k and contributors
# For license information, please see license.txt

import frappe
from frappe import _
from frappe.model.document import Document
from frappe.utils import cint, getdate

from ir.industrial_relations.utils import autoname_planning_document


class SitePlan(Document):
	def autoname(self):
		autoname_planning_document(self, "plan_name")

	def before_validate(self):
		self.ensure_group_keys()
		self.resolve_group_keys_by_label()
		self.remove_blank_child_rows()
		self.set_defaults()
		self.ensure_slot_keys()
		self.populate_display_values()

	def validate(self):
		self.validate_effective_dates()
		self.validate_group_keys_unique()
		self.validate_slots_reference_groups()
		self.validate_reporting_lines_reference_groups()

	def remove_blank_child_rows(self):
		self.groups = [row for row in self.groups or [] if _clean(row.group)]

		self.slots = [
			row
			for row in self.slots or []
			if _clean(row.group_key) and (_clean(row.designation) or _clean(row.asset_category) or _clean(row.row_label))
		]

		self.reporting_lines = [
			row
			for row in self.reporting_lines or []
			if _clean(row.source_group_key) and _clean(row.target_group_key)
		]

	def set_defaults(self):
		if not self.status:
			self.status = "Draft"

		if self.enabled is None:
			self.enabled = 1

	def ensure_group_keys(self):
		for row in self.groups or []:
			if not _clean(row.group_key):
				row.group_key = _new_group_key()

	def resolve_group_keys_by_label(self):
		"""A Slot/Reporting Line saved by anything other than the Designer's
		own UI - a raw Data Import in particular, since the exported CSV
		template has no group_key column at all, only the human-readable
		group label - only carries a group *label*. remove_blank_child_rows()
		right after this treats a missing group_key as "orphaned" and
		silently deletes the row, so this has to backfill the key from the
		label first, or every such row vanishes the moment the document is
		saved with no error raised at all.
		"""
		by_label = {_clean(row.group): row.group_key for row in self.groups or [] if _clean(row.group)}

		for row in self.slots or []:
			if not _clean(row.group_key):
				match = by_label.get(_clean(row.group))
				if match:
					row.group_key = match

		for row in self.reporting_lines or []:
			for prefix in ("source", "target"):
				key_field = f"{prefix}_group_key"
				label_field = f"{prefix}_group"
				if not _clean(getattr(row, key_field, None)):
					match = by_label.get(_clean(getattr(row, label_field, None)))
					if match:
						setattr(row, key_field, match)

	def ensure_slot_keys(self):
		# A blank row_key gets a fresh one, same as before - but a row_key
		# that *collides* with an earlier Slot's (seen in real imported data:
		# many Asset slots sharing the exact same row_key) needs exactly the
		# same repair, or every Slot after the first with that key silently
		# merges into one another downstream, wherever row_key is used as a
		# uniqueness key (Site Organogram population in particular).
		seen = set()
		for row in self.slots or []:
			key = _clean(row.row_key)
			if not key or key in seen:
				row.row_key = _new_slot_key()
			seen.add(_clean(row.row_key))

	def populate_display_values(self):
		group_names = {
			_clean(row.group_key): _clean(row.group)
			for row in self.groups or []
			if _clean(row.group_key)
		}

		for row in self.slots or []:
			row.group = group_names.get(_clean(row.group_key), row.group)

		for row in self.reporting_lines or []:
			row.source_group = group_names.get(_clean(row.source_group_key), row.source_group)
			row.target_group = group_names.get(_clean(row.target_group_key), row.target_group)

	def validate_effective_dates(self):
		if (
			self.effective_from
			and self.effective_until
			and getdate(self.effective_until) < getdate(self.effective_from)
		):
			frappe.throw(_("Effective Until cannot be before Effective From."))

	def validate_group_keys_unique(self):
		seen = set()
		for row in self.groups or []:
			key = _clean(row.group_key)
			if key in seen:
				frappe.throw(_("Duplicate group key for heading {0}.").format(row.group))
			seen.add(key)

	def validate_slots_reference_groups(self):
		valid_keys = {_clean(row.group_key) for row in self.groups or []}
		for row in self.slots or []:
			if _clean(row.group_key) not in valid_keys:
				frappe.throw(_("Slot {0} refers to a group heading that no longer exists.").format(row.row_label or row.idx))

	def validate_reporting_lines_reference_groups(self):
		valid_keys = {_clean(row.group_key) for row in self.groups or []}
		for row in self.reporting_lines or []:
			if _clean(row.source_group_key) not in valid_keys or _clean(row.target_group_key) not in valid_keys:
				frappe.throw(_("Reporting line {0} refers to a group heading that no longer exists.").format(row.idx))


def _clean(value):
	return str(value or "").strip()


def _new_group_key():
	return f"GRP::{frappe.generate_hash(length=10)}"


def _new_slot_key():
	return f"SLOT::{frappe.generate_hash(length=10)}"


def _team_count(shift_design):
	if not shift_design:
		return 0
	try:
		return max(0, int(frappe.db.get_value("Shift Design", shift_design, "number_of_teams") or 0))
	except (TypeError, ValueError):
		return 0


# -------------------------------------------------------------------
# Excel export
# -------------------------------------------------------------------

@frappe.whitelist()
def export_site_plan_excel(name):
	if not name:
		frappe.throw(_("Site Plan is required."))

	doc = frappe.get_doc("Site Plan", name)
	doc.check_permission("read")

	try:
		from openpyxl import Workbook
		from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
		from openpyxl.utils import get_column_letter
	except ImportError:
		frappe.throw(_("openpyxl is required for this export but is not installed."))

	from collections import defaultdict
	from io import BytesIO

	wb = Workbook()
	ws = wb.active
	ws.title = (doc.plan_name or "Site Plan")[:31]

	thin_side = Side(style="thin", color="000000")
	styles = {
		"title_font": Font(bold=True, size=14),
		"section_font": Font(bold=True, size=12),
		"header_font": Font(bold=True, size=10),
		"section_fill": PatternFill("solid", fgColor="D9EAD3"),
		"header_fill": PatternFill("solid", fgColor="E7E6E6"),
		"center": Alignment(horizontal="center", vertical="center", wrap_text=True),
		"wrap": Alignment(vertical="top", wrap_text=True),
		"thin_border": Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side),
	}

	row_no = 1
	ws.merge_cells(start_row=row_no, start_column=1, end_row=row_no, end_column=3)
	title_cell = ws.cell(row_no, 1, (doc.plan_name or doc.name or "SITE PLAN").upper())
	title_cell.font = styles["title_font"]
	title_cell.alignment = styles["center"]
	row_no += 1

	period = f"{doc.effective_from or ''} to {doc.effective_until or 'indefinite'}"
	ws.cell(row_no, 1, f"Effective: {period}")
	row_no += 1
	if doc.branch or doc.location:
		ws.cell(row_no, 1, f"Default Branch / Location: {doc.branch or '-'} / {doc.location or '-'}")
		row_no += 1
	row_no += 1

	groups = [row for row in doc.groups or [] if row.group]
	shift_counts = {row.group_key: _team_count(row.shift_design) for row in groups}

	row_no = _write_table(
		ws, row_no, "GROUP HEADINGS",
		["GROUP", "SHIFT DESIGN", "SHIFT TEAMS"],
		[[row.group, row.shift_design or "", shift_counts.get(row.group_key, 0)] for row in groups],
		styles,
	)

	# An Asset slot is *one* physical asset that gets a row in every shift
	# column of its group - a different Designation/Employee operates it in
	# each shift (rotation), but it's still the same single machine, so the
	# Asset Category total counts the Slot itself once, never multiplied by
	# the group's shift-team count. A Designation slot (or the Designation
	# attached to an Asset slot - the role that operates it) is the opposite:
	# a *different* person is needed for every shift the group runs, so that
	# total *is* multiplied by the shift-team count, the same expansion
	# get_site_plan_template() (ir.industrial_relations.doctype.site_organogram
	# .site_organogram) uses to build one vacant Organogram row per shift
	# column, kept independent here rather than importing across doctypes.
	# A Spare/Swing Asset is a backup unit that's never staffed, so it counts
	# toward the Asset total but never toward the Designation/headcount total
	# even if a Designation value happens to be set on it.
	designation_totals = defaultdict(int)
	category_totals = defaultdict(int)
	for slot in doc.slots or []:
		count = shift_counts.get(slot.group_key, 0)
		is_spare = cint(slot.spare_swing)

		if slot.row_type == "Designation" and slot.designation:
			designation_totals[slot.designation] += count
		if slot.row_type == "Asset":
			if slot.asset_category:
				category_totals[slot.asset_category] += 1
			if slot.designation and not is_spare:
				designation_totals[slot.designation] += count

	row_no = _write_table(
		ws, row_no, "TOTAL EMPLOYEES REQUIRED BY DESIGNATION",
		["DESIGNATION", "REQUIRED COUNT"],
		sorted(([label, count] for label, count in designation_totals.items()), key=lambda r: r[0]),
		styles,
	)

	row_no = _write_table(
		ws, row_no, "TOTAL ASSETS REQUIRED BY ASSET CATEGORY",
		["ASSET CATEGORY", "REQUIRED COUNT"],
		sorted(([label, count] for label, count in category_totals.items()), key=lambda r: r[0]),
		styles,
	)

	for col_idx in range(1, ws.max_column + 1):
		letter = get_column_letter(col_idx)
		ws.column_dimensions[letter].width = 28 if col_idx == 1 else 20

	out = BytesIO()
	wb.save(out)
	out.seek(0)

	filename = f"{frappe.scrub(doc.name or 'site_plan')}.xlsx"
	frappe.local.response.filename = filename
	frappe.local.response.filecontent = out.getvalue()
	frappe.local.response.type = "binary"


def _write_table(ws, row_no, title, headers, rows, styles):
	total_cols = max(1, len(headers))

	ws.merge_cells(start_row=row_no, start_column=1, end_row=row_no, end_column=total_cols)
	title_cell = ws.cell(row_no, 1, title)
	title_cell.font = styles["section_font"]
	title_cell.alignment = styles["center"]
	title_cell.fill = styles["section_fill"]
	row_no += 1

	for idx, header in enumerate(headers, start=1):
		cell = ws.cell(row_no, idx, header)
		cell.font = styles["header_font"]
		cell.fill = styles["header_fill"]
		cell.alignment = styles["center"]
	row_no += 1
	data_start = row_no

	if rows:
		for item in rows:
			for idx, value in enumerate(item, start=1):
				cell = ws.cell(row_no, idx, value)
				cell.alignment = styles["wrap"]
			row_no += 1
	else:
		ws.cell(row_no, 1, "None")
		row_no += 1

	for row in ws.iter_rows(min_row=data_start - 2, max_row=row_no - 1, min_col=1, max_col=total_cols):
		for cell in row:
			cell.border = styles["thin_border"]

	return row_no + 1
