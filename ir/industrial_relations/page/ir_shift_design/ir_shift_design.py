# Copyright (c) 2026, BuFf0k and contributors
# For license information, please see license.txt

from __future__ import annotations

import json
from datetime import datetime

import frappe
from frappe import _
from frappe.utils import cint, getdate


SHIFT_DESIGN = "Shift Design"
TABLE_FIELDS = (
	"shift_types",
	"teams",
	"pattern",
	"calendar_rules",
	"date_overrides",
)


@frappe.whitelist()
def get_bootstrap():
	_check_permission("read")

	return {
		"designs": list_designs(),
		"companies": _names("Company", group_filter=True),
		"shift_types": _shift_type_options(),
		"can_create": frappe.has_permission(
			SHIFT_DESIGN,
			ptype="create",
		),
		"parent_fields": _fieldnames(SHIFT_DESIGN),
		"shift_type_fields": _child_fieldnames(SHIFT_DESIGN, "shift_types"),
		"team_fields": _child_fieldnames(SHIFT_DESIGN, "teams"),
		"pattern_fields": _child_fieldnames(SHIFT_DESIGN, "pattern"),
		"calendar_rule_fields": _child_fieldnames(
			SHIFT_DESIGN,
			"calendar_rules",
		),
		"date_override_fields": _child_fieldnames(
			SHIFT_DESIGN,
			"date_overrides",
		),
	}


@frappe.whitelist()
def list_designs():
	_check_permission("read")
	meta = frappe.get_meta(SHIFT_DESIGN)

	candidate_fields = (
		"name",
		"design_name",
		"branch",
		"company",
		"status",
		"enabled",
		"effective_from",
		"effective_until",
		"number_of_teams",
		"cycle_length",
		"anchor_date",
		"pay_period_start_day",
		"pay_period_end_day",
		"modified",
	)

	fields = [
		fieldname
		for fieldname in candidate_fields
		if fieldname in {"name", "modified"}
		or meta.has_field(fieldname)
	]

	return frappe.get_all(
		SHIFT_DESIGN,
		fields=fields,
		order_by="modified desc",
		limit_page_length=500,
	)


@frappe.whitelist()
def get_design(name):
	if not name:
		frappe.throw(_("Shift Design is required."))

	doc = frappe.get_doc(SHIFT_DESIGN, name)
	doc.check_permission("read")
	return _serialize(doc)


@frappe.whitelist()
def save_design(data):
	payload = _json_object(data)
	name = _clean(payload.get("name"))

	if name:
		doc = frappe.get_doc(SHIFT_DESIGN, name)
		doc.check_permission("write")
	else:
		if not frappe.has_permission(SHIFT_DESIGN, ptype="create"):
			frappe.throw(
				_("You do not have permission to create Shift Designs."),
				frappe.PermissionError,
			)
		doc = frappe.new_doc(SHIFT_DESIGN)

	meta = frappe.get_meta(SHIFT_DESIGN)

	for fieldname, value in payload.items():
		if fieldname in TABLE_FIELDS or fieldname == "name":
			continue
		if meta.has_field(fieldname):
			doc.set(fieldname, value)

	for table_fieldname in TABLE_FIELDS:
		table_field = meta.get_field(table_fieldname)
		if not table_field or not table_field.options:
			continue

		doc.set(table_fieldname, [])
		for row in payload.get(table_fieldname) or []:
			if not isinstance(row, dict):
				continue
			doc.append(
				table_fieldname,
				_clean_child_payload(row, table_field.options),
			)

	if doc.is_new():
		doc.insert()
	else:
		doc.save()

	return {
		"design": _serialize(doc),
		"designs": list_designs(),
	}


@frappe.whitelist()
def delete_design(name):
	if not name:
		frappe.throw(_("Shift Design is required."))

	doc = frappe.get_doc(SHIFT_DESIGN, name)
	doc.check_permission("delete")
	frappe.delete_doc(SHIFT_DESIGN, name)

	return {"designs": list_designs()}


@frappe.whitelist()
def get_sa_public_holidays(start_date, end_date):
	if not start_date or not end_date:
		return []

	start = getdate(start_date)
	end = getdate(end_date)

	if end < start:
		frappe.throw(_("Simulation End cannot be before Simulation Start."))

	years = list(range(start.year, end.year + 1))

	try:
		from holidays import country_holidays

		za_holidays = country_holidays("ZA", years=years)
		return [
			{
				"date": str(getdate(holiday_date)),
				"description": holiday_name,
			}
			for holiday_date, holiday_name in za_holidays.items()
			if start <= getdate(holiday_date) <= end
		]
	except Exception as exc:
		frappe.log_error(
			title="Shift Designer public holiday generation failed",
			message=frappe.get_traceback(),
		)
		frappe.throw(
			_(
				"Unable to generate South African public holidays. "
				"Please confirm that the Python 'holidays' package is installed. "
				"Original error: {0}"
			).format(exc)
		)


@frappe.whitelist()
def export_shift_design_excel(name):
	if not name:
		frappe.throw(_("Shift Design is required."))

	doc = frappe.get_doc(SHIFT_DESIGN, name)
	doc.check_permission("read")

	try:
		from openpyxl import Workbook
		from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
		from openpyxl.utils import get_column_letter
	except ImportError:
		frappe.throw(_("openpyxl is required for this export but is not installed."))

	from io import BytesIO

	wb = Workbook()
	ws = wb.active
	ws.title = (doc.design_name or "Shift Design")[:31]

	thin_side = Side(style="thin", color="000000")
	styles = {
		"title_font": Font(bold=True, size=14),
		"section_font": Font(bold=True, size=12),
		"header_font": Font(bold=True, size=10),
		"section_fill": PatternFill("solid", fgColor="D9EAD3"),
		"header_fill": PatternFill("solid", fgColor="E7E6E6"),
		"center": Alignment(horizontal="center", vertical="center", wrap_text=True),
		"wrap": Alignment(vertical="top", wrap_text=True),
		"thin_border": Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side),
	}

	row_no = 1
	ws.merge_cells(start_row=row_no, start_column=1, end_row=row_no, end_column=4)
	title_cell = ws.cell(row_no, 1, (doc.design_name or doc.name or "SHIFT DESIGN").upper())
	title_cell.font = styles["title_font"]
	title_cell.alignment = styles["center"]
	row_no += 1

	period = f"{doc.effective_from or ''} to {doc.effective_until or 'indefinite'}"
	ws.cell(row_no, 1, f"Effective: {period}  |  Status: {doc.status or ''}  |  Branch: {doc.branch or '-'}")
	row_no += 1
	ws.cell(
		row_no, 1,
		f"Cycle Length: {doc.cycle_length}  |  Anchor Date: {doc.anchor_date or '-'}  |  "
		f"Pay Period: day {doc.pay_period_start_day} to day {doc.pay_period_end_day}  |  "
		f"Ordinary Hours Limit: {doc.ordinary_hours_limit}",
	)
	row_no += 2

	teams = sorted(
		[row for row in doc.teams or [] if cint(row.enabled)],
		key=lambda row: cint(row.display_order),
	)

	# Shift Type is a Link - the actual start/end time and colour live on the
	# real "Shift Type" doctype record, not on this Design's own child row.
	shift_type_names = [row.shift_type for row in doc.shift_types or [] if row.shift_type]
	shift_type_info = {}
	if shift_type_names:
		shift_type_info = {
			row.name: row
			for row in frappe.get_all(
				"Shift Type",
				filters={"name": ["in", shift_type_names]},
				fields=["name", "start_time", "end_time", "color"],
			)
		}

	row_no = _write_table(
		ws, row_no, "SHIFT TYPES",
		["SHIFT TYPE", "START", "END", "HOURS", "COLOUR"],
		[
			[
				row.shift_type,
				str(shift_type_info.get(row.shift_type, {}).get("start_time") or ""),
				str(shift_type_info.get(row.shift_type, {}).get("end_time") or ""),
				_duration_hours(
					shift_type_info.get(row.shift_type, {}).get("start_time"),
					shift_type_info.get(row.shift_type, {}).get("end_time"),
				),
				shift_type_info.get(row.shift_type, {}).get("color") or "",
			]
			for row in doc.shift_types or []
		],
		styles,
	)

	row_no = _write_table(
		ws, row_no, "SHIFT TEAMS",
		["TEAM", "DISPLAY ORDER", "PATTERN OFFSET"],
		[[row.team_name or row.team_key, row.display_order, row.pattern_offset] for row in teams],
		styles,
	)

	pattern_by_team_day = {
		(row.team_key, cint(row.pattern_day)): row.assignment or ""
		for row in doc.pattern or []
	}
	cycle_length = max(cint(doc.cycle_length), 1)
	pattern_rows = [
		[f"Day {day}"] + [pattern_by_team_day.get((team.team_key, day), "") or "Off" for team in teams]
		for day in range(1, cycle_length + 1)
	]
	row_no = _write_table(
		ws, row_no, "ROTATION PATTERN",
		["CYCLE DAY"] + [team.team_name or team.team_key for team in teams],
		pattern_rows,
		styles,
	)

	row_no = _write_table(
		ws, row_no, "CALENDAR RULES",
		["RULE TYPE", "DAY OF WEEK", "ACTION", "TARGET SHIFT TYPE", "HOURS OVERRIDE", "PRIORITY", "ENABLED"],
		[
			[
				row.rule_type, row.day_of_week or "", row.action,
				row.target_shift_type or "", row.hours_override or "",
				row.priority, "Yes" if cint(row.enabled if row.enabled is not None else 1) else "No",
			]
			for row in doc.calendar_rules or []
		],
		styles,
	)

	date_overrides = doc.date_overrides or []
	if date_overrides:
		row_no = _write_table(
			ws, row_no, "DATE OVERRIDES",
			["DATE", "TEAM", "ASSIGNMENT", "REASON"],
			[
				[
					str(row.date) if row.date else "",
					next((t.team_name or t.team_key for t in teams if t.team_key == row.team_key), row.team_key or ""),
					row.assignment or "Off",
					row.reason or "",
				]
				for row in date_overrides
			],
			styles,
		)

	for col_idx in range(1, ws.max_column + 1):
		letter = get_column_letter(col_idx)
		ws.column_dimensions[letter].width = 26 if col_idx == 1 else 16

	out = BytesIO()
	wb.save(out)
	out.seek(0)

	filename = f"{frappe.scrub(doc.name or 'shift_design')}.xlsx"
	frappe.local.response.filename = filename
	frappe.local.response.filecontent = out.getvalue()
	frappe.local.response.type = "binary"


def _write_table(ws, row_no, title, headers, rows, styles):
	total_cols = max(1, len(headers))

	ws.merge_cells(start_row=row_no, start_column=1, end_row=row_no, end_column=total_cols)
	title_cell = ws.cell(row_no, 1, title)
	title_cell.font = styles["section_font"]
	title_cell.alignment = styles["center"]
	title_cell.fill = styles["section_fill"]
	row_no += 1

	for idx, header in enumerate(headers, start=1):
		cell = ws.cell(row_no, idx, header)
		cell.font = styles["header_font"]
		cell.fill = styles["header_fill"]
		cell.alignment = styles["center"]
	row_no += 1
	data_start = row_no

	if rows:
		for item in rows:
			for idx, value in enumerate(item, start=1):
				cell = ws.cell(row_no, idx, value)
				cell.alignment = styles["wrap"]
			row_no += 1
	else:
		ws.cell(row_no, 1, "None")
		row_no += 1

	for row in ws.iter_rows(min_row=data_start - 2, max_row=row_no - 1, min_col=1, max_col=total_cols):
		for cell in row:
			cell.border = styles["thin_border"]

	return row_no + 1


def _serialize(doc):
	meta = frappe.get_meta(SHIFT_DESIGN)
	data = {"name": doc.name}

	for field in meta.fields:
		if field.fieldtype in {
			"Section Break",
			"Column Break",
			"Tab Break",
			"HTML",
			"Button",
		}:
			continue

		if field.fieldtype == "Table":
			data[field.fieldname] = [
				_serialize_child(row)
				for row in doc.get(field.fieldname) or []
			]
		else:
			data[field.fieldname] = doc.get(field.fieldname)

	data["modified"] = doc.modified
	return data


def _serialize_child(row):
	meta = frappe.get_meta(row.doctype)
	return {
		field.fieldname: row.get(field.fieldname)
		for field in meta.fields
		if field.fieldtype not in {
			"Section Break",
			"Column Break",
			"Tab Break",
			"HTML",
			"Button",
		}
	}


def _clean_child_payload(row, child_doctype):
	valid_fields = set(_fieldnames(child_doctype))
	return {
		key: value
		for key, value in row.items()
		if key in valid_fields
	}


def _shift_type_options():
	if not frappe.db.exists("DocType", "Shift Type"):
		return []

	meta = frappe.get_meta("Shift Type")
	fields = ["name"]
	for fieldname in ("start_time", "end_time", "color"):
		if meta.has_field(fieldname):
			fields.append(fieldname)

	rows = frappe.get_all(
		"Shift Type",
		filters=_active_filters(meta),
		fields=fields,
		order_by="name asc",
	)

	return [
		{
			"name": row.name,
			"start_time": str(row.get("start_time") or ""),
			"end_time": str(row.get("end_time") or ""),
			"color": row.get("color") or "",
			"hours": _duration_hours(
				row.get("start_time"),
				row.get("end_time"),
			),
		}
		for row in rows
	]


def _duration_hours(start, end):
	if start is None or end is None:
		return 0

	def seconds(value):
		if hasattr(value, "total_seconds"):
			return value.total_seconds()

		text = str(value).split(".")[0]
		for date_format in ("%H:%M:%S", "%H:%M"):
			try:
				parsed = datetime.strptime(text, date_format)
				return (
					parsed.hour * 3600
					+ parsed.minute * 60
					+ parsed.second
				)
			except ValueError:
				continue
		return 0

	start_seconds = seconds(start)
	end_seconds = seconds(end)
	if end_seconds <= start_seconds:
		end_seconds += 24 * 3600

	return round((end_seconds - start_seconds) / 3600, 4)


def _names(doctype, group_filter=False):
	if not frappe.db.exists("DocType", doctype):
		return []

	meta = frappe.get_meta(doctype)
	filters = {}

	if group_filter and meta.has_field("is_group"):
		filters["is_group"] = 0
	if meta.has_field("disabled"):
		filters["disabled"] = 0

	return frappe.get_all(
		doctype,
		filters=filters,
		pluck="name",
		order_by="name asc",
	)


def _active_filters(meta):
	if meta.has_field("disabled"):
		return {"disabled": 0}
	if meta.has_field("enabled"):
		return {"enabled": 1}
	return {}


def _fieldnames(doctype):
	return [
		field.fieldname
		for field in frappe.get_meta(doctype).fields
		if field.fieldtype not in {
			"Section Break",
			"Column Break",
			"Tab Break",
			"HTML",
			"Button",
		}
	]


def _child_fieldnames(parent_doctype, table_fieldname):
	parent_meta = frappe.get_meta(parent_doctype)
	table_field = parent_meta.get_field(table_fieldname)
	if not table_field or not table_field.options:
		return []
	return _fieldnames(table_field.options)


def _json_object(value):
	if isinstance(value, dict):
		return value

	try:
		parsed = json.loads(value or "{}")
	except (TypeError, ValueError):
		frappe.throw(_("Invalid Shift Design payload."))

	if not isinstance(parsed, dict):
		frappe.throw(_("Shift Design payload must be an object."))

	return parsed


def _check_permission(ptype):
	if not frappe.has_permission(SHIFT_DESIGN, ptype=ptype):
		frappe.throw(_("Not permitted."), frappe.PermissionError)


def _clean(value):
	return str(value or "").strip()
